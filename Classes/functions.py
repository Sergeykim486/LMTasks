import os, folium, openpyxl, re, io, Classes.buttons as buttons
from folium.plugins import MarkerCluster
from jinja2 import Template
from folium import IFrame
from urllib.parse import quote
from Classes.db import Database
from datetime import datetime
from openpyxl.styles import Alignment
from Classes.config import bot, sendedmessages, db
day_translation = {
    'Monday': 'Понедельник',
    'Tuesday': 'Вторник',
    'Wednesday': 'Среда',
    'Thursday': 'Четверг',
    'Friday': 'Пятница',
    'Saturday': 'Суббота',
    'Sunday': 'Воскресенье'
}


dbname = os.path.dirname(os.path.abspath(__file__)) + '/Database/' + 'lmtasksbase.db'
db = Database(dbname)

def sendtoall(message, markdown, exeptions, nt = 0, notific = False):
    global sendedmessages
    users = db.select_table('Users')
    for user in users:
        # logging.info(f'\nℹ️ sended message to user {user[2]} {user[1]}')
        try:
            if user[0] != exeptions:
                mes = bot.send_message(
                    user[0],
                    message,
                    reply_markup=markdown,
                    disable_notification=notific
                )
                if nt == 1:
                    sendedmessages.append([[user[0]], [mes.message_id]])
        except Exception as e:
            # logging.info('Пользователь заблокировал бота...')
            pass
    return

def mesdel(chat_id, mess_id):
    try:
        bot.delete_message(chat_id, mess_id)
    except Exception as e:
        pass

def conftext(message, ActiveUser):
    mestext = "Подтвердите правильность введенной информации.\n"
    mestext = mestext + "Ваш id:" + str(ActiveUser[message.chat.id]["id"]) + "\n"
    mestext = mestext + "Имя:" + str(ActiveUser[message.chat.id]["FirstName"]) + "\n"
    mestext = mestext + "Фамилия:" + str(ActiveUser[message.chat.id]["LastName"]) + "\n"
    mestext = mestext + "Телефон:" + str(ActiveUser[message.chat.id]["PhoneNumber"]) + "\n"
    return mestext

def listgen(table, cols, tasks = 0):
    res = []
    for line in table:
        curline = ''
        for i in cols:

            if tasks == 1:

                if i == 0:
                    curline = curline + '№ ' + (str(line[i]) if line[i] is not None else '-') + '\n'

                elif i == 1:
                    curline = curline + ' от ' + (str(line[i]) if line[i] is not None else '') + '\n'

                elif i == 3:
                    loc = db.get_record_by_id('Locations', line[12])
                    if loc == None:
                        location = ''
                    else:
                        location = loc[2]
                    cname = str(db.get_record_by_id('Contragents', line[i])[0]) + " " + str(db.get_record_by_id('Contragents', line[i])[1]) + "\n Локация - " + location + "\nДоговор: " + str(db.get_record_by_id('Contragents', line[i])[6]) + '\n\n' + str(db.get_record_by_id('Contragents', line[i])[3]) + ': ' + str(db.get_record_by_id('Contragents', line[i])[4])
                    curline = curline + str(cname) + '\n'

                elif i == 6 and line[i]:
                    if line [i] is None:
                        Mastername = '-'
                    else:
                        try:
                            user_record = db.get_record_by_id("Users", line[i])
                            if user_record is not None:
                                Mastername = str(user_record[2]) + ' ' + str(user_record[1])
                            else:
                                Mastername = "-"
                        except TypeError:
                            Mastername = "-"
                    curline = curline + '\nМастер: - ' + str(Mastername) + '\n'

                else:

                    if i != 6:
                        curline = curline + str(line[i]) + '\n'

            else:
                curline = curline + str(line[i]) + ' '

        if tasks == 0:
            res.append(curline)

        elif tasks == 1:
            marker = ''

            if str(line[11]) == '1': marker = '🔵 '

            elif str(line[11]) == '2': marker = '🟡 '

            elif str(line[11]) == '3': marker = '🟢 '

            elif str(line[11]) == '4': marker = '🔴 '

            elif str(line[11]) == '5': marker = '🟦 '
            
            elif str(line[11]) == '6': marker = '🟨 '
            
            elif str(line[11]) == '7': marker = '🟩 '
            
            else: marker = '⚪️ '

            res.append(marker + curline)

        elif tasks == 2:
            res.append('🗄 ' + curline)

        elif tasks == 3:
            res.append('👤 ' + curline)

    return res

def curtask(id):
    messtext = ''
    task = db.get_record_by_id('Tasks', id)
    messtext = 'Заявка №' + str(task[0]) + ' от ' + str(task[1])
    messtext = messtext + '\nЗаявку зарегистрировал(а): ' + str(db.get_record_by_id('Users', task[2])[2]) + ' ' + str(db.get_record_by_id('Users', task[2])[1])

    if task[11] == 2 or task[10] == 3:
        
        try:
            user_record = db.get_record_by_id('Users', task[6])
            if user_record is not None:
                master = str(user_record[2]) + ' ' + str(user_record[1])
            else:
                master = "-"
        except TypeError:
            master = "-"
        messtext = messtext + '\nМастер ' + master
        messtext = messtext + ' принял ' + str(task[5])

    elif task[11] == 4:
        messtext = messtext + '\n' + str(db.get_record_by_id('Users', task[9])[1]) + ' отменил заявку\nПРИЧИНА ОТМЕНЫ:\n' + str(task[10])

    if task[10] is not None:
        messtext = messtext + '\n❗️ ' + str(task[10])
    loc = db.get_record_by_id('Locations', task[12])
    if loc == None:
        location = ''
    else:
        location = loc[2]
    messtext = messtext + '\n\nПоступила от: ' + str(db.get_record_by_id('Contragents', task[3])[0]) + " " + str(db.get_record_by_id('Contragents', task[3])[1]) + " \nЛокация - " + location + "\nДоговор: " +str(db.get_record_by_id('Contragents', task[3])[6])
    messtext = messtext + '\n\nЗАПРОС:\n' + str(task[4])
    messtext = messtext + '\n\nКОНТАКТЫ ЗАКАЗЧИКА:\n' + str(db.get_record_by_id('Contragents', task[3])[3]) + ' - ' + str(db.get_record_by_id('Contragents', task[3])[4])
    messtext = messtext + f'\nАдрес: ' + str(db.get_record_by_id('Contragents', task[3])[2])
    messtext = messtext + '\n\nСтатус заявки - ' + str(db.get_record_by_id('Statuses', task[11])[1])
    return messtext

def getuserlist():
    users = db.select_table('Users')
    userlist = []
    for line in users:
        userlist.append(line[0])
    return userlist

# Добавление счетчика в частые
def top10add(client, user):
    Top = db.select_table_with_filters("Top10", {'uid': user})
    finded = False
    for i in Top:
        if i[2] == client[0]:
            quantity = db.get_record_by_id('Top10', i[0])[3] + 1
            db.update_records(
                'Top10',
                ['val'],
                [quantity],
                "id",
                i[0]
            )
            finded = True
    if finded == False:
        db.insert_record(
            "Top10",
            [
                None,
                user,
                client[0],
                1
            ]
        )
def top10buttons(user):
    data = db.select_table_with_filters('Top10', {'uid': user})
    sorted_data = sorted(data, key=lambda x: x[3], reverse=True)
    top_10 = sorted_data[:5] if len(sorted_data) >= 5 else sorted_data
    buttonscont = []
    buttonscont.append('🚫 Отмена')
    for contr in top_10:
        cont = db.get_record_by_id('Contragents', contr[2])
        line = str(cont[0]) + ' ' + str(cont[1])
        buttonscont.append(line)
    return buttonscont


# Карта заявок

def mmapgen(locations):
    # Создание объекта карты
    map = folium.Map(location=[41.28927613679946, 69.31295641163192], zoom_start=12)

    # Создание группы маркеров
    marker_cluster = MarkerCluster().add_to(map)

    # Создание списка точек с попапами
    popup_content = ""

    # Легенды
    legend_content = '''
        <div style="font-size: 20px; color: black;"><b>Условные обозначения:</div><br/>
        <div style="margin-bottom: 10px; color: blue;"><b style="color: blue;">████</b> - Заявки не принятые мастерами</b></div>
        <div style="margin-bottom: 10px; color: orange;"><b style="color: orange;">████</b> - Заявки у мастеров</b></div>
        <div style="margin-bottom: 10px; color: green;"><b style="color: green;">████</b> - Выполненные заявки</b></div>
        <div style="margin-bottom: 10px; color: red;"><b style="color: red;">████</b> - Отмененные заявки</div><br/>
    '''
    
    legend_popup = folium.Popup(IFrame(html=legend_content, width=200, height=150), max_width=200)
    legend_marker = folium.Marker(location=[41.28921489333344, 69.31288111459628], icon=folium.Icon(color='gray', icon='info-sign'), popup=legend_popup)
    legend_marker.add_to(map)

    for location in locations:
        name = location[0].replace('|', '<br>')
        taskid = name.split()[1]
        description = location[1].replace('|', '<br>')
        lat = float(location[2])
        lon = float(location[3])
        color = 'blue'
        if location[4] == 2:
            color = 'orange'
        elif location[4] == 3:
            color = 'green'
        elif location[4] == 4:
            color = 'red'
        link = f'https://t.me/labmonotasktelebot?start={taskid}'
        point_content = f'<div style="font-size: 16px; color: #0057B5;"><b>{name}</b></div><div style="font-size: 12px; color: black;">{description}</div><div><button style="margin-top: 10px;" onclick="window.open(\'{link}\', \'_blank\')">Перейти к заявке</button></div><br/>'
        point_popup = folium.Popup(IFrame(html=point_content, width=300, height=240), max_width=300)
        folium.Marker([lat, lon], popup=point_popup, icon=folium.Icon(color=color)).add_to(marker_cluster)
        
        # Добавление точки в список попапов
        # popup_content += f'<div style="margin-bottom: 10px; color: {color};"><b style="color: {color};">{name}</b><br/><span style="color: black;">{description}</span><br/><button onclick="window.open(\'{link}\', \'_blank\')">Перейти к заявке</button><br/><br/></div>'
        popup_content += f'<div style="margin-bottom: 10px; color: {color};"><b style="color: {color};">{name}</b><br/><span style="color: black; width: 280px;">{description}</span><br/><button onclick="window.open(\'{link}\', \'_blank\')">Перейти к заявке</button><br/><br/></div>'


    # Добавление легенды в список попапов
    popup_content = legend_content + popup_content

    # Создание маркера с иконкой информации
    legend_marker = folium.Marker(location=[41.28921489333344, 69.31288111459628], icon=folium.Icon(color='gray', icon='info-sign'))
    folium.Popup(IFrame(html=popup_content, width=300, height=600), max_width=300).add_to(legend_marker)
    legend_marker.add_to(map)

    # Добавление кнопки обновления страницы
    refresh_button = '''
        <button onclick="location.reload();" style="position: absolute; top: 10px; right: 10px; z-index: 9999; background-color: green; color: white;">Обновить страницу</button>
    '''
    map.get_root().html.add_child(folium.Element(refresh_button))

    # Сохранение карты в HTML-файл
    map.save('public/m.map.html')

# =======================================================================================================

def generate_popup(location):
    name = location[0].replace('|', '<br>')
    description = location[1].replace('|', '<br>')
    status = location[4]
    task_id = name.split()[1]
    link = f"https://t.me/labmonotasktelebot?start={task_id}"

    # Set color based on status
    color = 'black'
    if status == 1:
        color = 'blue'
    elif status == 2:
        color = 'orange'
    elif status == 3:
        color = 'green'
    elif status == 4:
        color = 'red'

    popup_html = f"""
        <div style="width: auto;">
            <p style="color:{color}; font-size:16px; font-weight:bold;">{name}</p>
            <p style="color:black; font-size:12px;">{description}</p>
            <button onclick="window.open('{link}', '_blank')" style="background-color:{color}; color:white; font-size:16px;">Перейти к заявке</button>
        </div>
    """
    return popup_html

def generate_marker_html(location):
    name = location[0].replace('|', '<br>')
    description = location[1].replace('|', '<br>')
    status = location[4]

    color = 'black'
    if status == 1:
        color = 'blue'
    elif status == 2:
        color = 'orange'
    elif status == 3:
        color = 'green'
    elif status == 4:
        color = 'red'

    marker_html = f"""
        <div>
            <p style="color:{color}; font-size:16px; font-weight:bold;">{name}</p>
            <p style="color:black; font-size:12px;">{description}</p>
            <button onclick="window.open('https://t.me/labmonotasktelebot?start={name.split()[1]}', '_blank')" style="background-color:{color}; color:white; font-size:16px;">Перейти к заявке</button><br><br>
        </div>
    """
    return marker_html

def mapgen(locations):
    # Create map
    map_object = folium.Map(location=[41.28927613679946, 69.31295641163192], zoom_start=12)

    # Create marker cluster
    marker_cluster = MarkerCluster().add_to(map_object)

    # Add markers to the cluster
    for location in locations:
        name = location[0]
        lat = location[2]
        lon = location[3]
        status = location[4]

        # Create popup for the marker
        popup_html = generate_popup(location)
        popup = folium.Popup(popup_html, max_width='auto', style='background-color: rgba(255, 255, 200, 0.9);')

        # Set marker color based on status
        marker_color = 'black'
        if status == 1:
            marker_color = 'blue'
        elif status == 2:
            marker_color = 'orange'
        elif status == 3:
            marker_color = 'green'
        elif status == 4:
            marker_color = 'red'

        # Create marker with popup
        marker = folium.Marker(location=[lat, lon], popup=popup, icon=folium.Icon(color=marker_color))
        marker.add_to(marker_cluster)

    # Generate HTML and save to file
    template = Template('''
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Активные заявки</title>
            <style>
                .button {
                    padding: 10px 20px;
                    display: inline-block;
                    border-radius: 4px;
                    font-size: 12px;
                    justify-self: flex-start;
                }
                .blue {
                    color: blue;
                }
                .orange {
                    color: orange;
                }
                .green {
                    color: green;
                }
                .red {
                    color: red;
                }
                .map-container {
                    position: absolute;
                    top: 100px;
                    left: 0;
                    right: 0;
                    bottom: 0;
                    overflow: hidden;
                }
                .popup-list {
                    position: absolute;
                    top: 110px;
                    left: 10px;
                    width: 30%;
                    height: calc(100% - 130px);
                    background-color: rgba(255, 255, 255, 0.9);
                    border: 1px solid black;
                    padding: 10px;
                    overflow: auto;
                }
                header {
                    background-color: lightgray;
                    height: 100px;
                    display: flex;
                    align-items: flex-start;
                    justify-content: space-between;
                }
                .logo {
                    margin-top: 20px;
                    margin-right: 20px;
                    align-self: flex-start;
                    justify-self: flex-end;
                }
            </style>
        </head>
        <body>
            <header>
                <div style="align-self: flex-start;">
                    <h1 style="margin-top: 0; margin-left: 20px;">Активные заявки</h1>
                    <button onclick="location.reload();" style="background-color: green; color: white; font-weight: bold; font-size: 18px; margin-top: -10px; margin-left: 20px;">Обновить страницу</button>
                    <button onclick="toggleList();" style="background-color: orange; color: white; font-weight: bold; font-size: 18px; margin-top: -10px; margin-left: 20px;">Показать список</button>
                </div>
                <div class="logo">
                    <img src="Logo.png" alt="Logo" style="height: 56px;">
                </div>
            </header>
            <div class="map-container">
                {{ folium_map|safe }}
            </div>
            <div id="popup-list" class="popup-list" style="display: none;">
                <h2 style="color: black; font-size: 18px; font-weight: bold;">Условные обозначения</h2>
                <p style="color: blue; font-size: 14px;">████ - Заявки не принятые мастерами</p>
                <p style="color: orange; font-size: 14px;">████ - Заявки у мастеров</p>
                <p style="color: green; font-size: 14px;">████ - Выполненные заявки</p>
                <p style="color: red; font-size: 14px;">████ - Отмененные заявки</p>
                <br>
                <div class="scrollable-list">
                    {% for location in locations %}
                        <div style="margin-bottom: 20px;">
                            {{ generate_marker_html(location) }}
                        </div>
                    {% endfor %}
                </div>
            </div>

            <script type="text/javascript">
                function toggleList() {
                    var list = document.getElementById("popup-list");
                    if (list.style.display === "none") {
                        list.style.display = "block";
                    } else {
                        list.style.display = "none";
                    }
                }
            </script>
            <script type="text/javascript">
                var isMobile = /iPhone|iPad|iPod|Android|webOS|BlackBerry|IEMobile|Opera Mini/i.test(navigator.userAgent);
                if (isMobile) {
                    window.location.href = 'm.map.html';
                }
            </script>
        </body>
        </html>
    ''')
    folium_map = map_object._repr_html_()
    html = template.render(folium_map=folium_map, locations=locations, generate_marker_html=generate_marker_html)

    with io.open('public/map.html', 'w', encoding='utf-8') as file:
        file.write(html)

# ======================================================================================================
# ПОИСК ЭЛЕМЕНТА ПО КЛЮЧЕВОМУ СЛОВУ
# ======================================================================================================

def transliterate_cyrillic_to_latin(word):
    conversion = {
        'кс': ['x'], 'а': ['a', 'u'], 'б': ['b'], 'в': ['v'], 'г': ['g'], 'д': ['d'], 'е': ['e', 'a'], 'ё': ['e'], 'ж': ['zh', 'j', 'dj'],
        'з': ['z'], 'и': ['i'], 'й': ['y', 'i', 'j'], 'к': ['k', 'q'], 'л': ['l'], 'м': ['m'], 'н': ['n'], 'о': ['o'],
        'п': ['p'], 'р': ['r'], 'с': ['s', 'c', 'ce', 'se'], 'т': ['t'], 'у': ['u', "o'"], 'ф': ['f'], 'х': ['kh', 'h', 'x'],
        'ц': ['ts', 'c'], 'ч': ['ch'], 'ш': ['sh'], 'щ': ['shch'], 'ъ': [''], 'ы': ['y'], 'ь': [''],
        'э': ['e', 'a'], 'ю': ['yu', 'iu'], 'я': ['ya', 'ia'],
        'a': ['а', 'э', 'е'], 'b': ['б'], 'c': ['ц', 'с', 'к'], 'd': ['д'], 'e': ['е', 'ё', 'и'], 'f': ['ф'], 'g': ['г', 'дж'], 'h': ['х'],
        'i': ['и', 'й', 'ай'], 'j': ['ж', 'дж'], 'k': ['к'], 'l': ['л'], 'm': ['м'], 'n': ['н'], 'o': ['о'], 'p': ['п'],
        'q': ['к'], 'r': ['р'], 's': ['с'], 't': ['т'], 'u': ['у', 'ю'], 'v': ['в'], 'w': ['в'], 'x': ['х', 'кс'], 'y': ['ы'],
        'z': ['з'], 'ce': ['с'], 'se': ['с'], 'ya': ['я'], 'yu': ['ю'], 'yo': ['ё'], 'ia': ['я'], 'iu': ['ю'], 'io': ['ё'], "o'": ['у']
    }

    if word.isalpha():
        result = ['']
        for letter in word:
            if letter in conversion:
                converted = conversion[letter]
                temp_result = []
                for w in result:
                    for char in converted:
                        temp_result.append(w + char)
                result = temp_result
            elif letter == 'й':
                converted = conversion[letter]
                temp_result = []
                for w in result:
                    for char in converted:
                        temp_result.append(w + char)
                        if char == 'i':
                            temp_result.append(w + 'y')
                        elif char == 'j':
                            temp_result.append(w + 'y')
                            temp_result.append(w + 'i')
                result = temp_result
        return result
    else:
        return []

def latin_change_liters(value1):
    liters = ['h kh', 'h x', 'x h', 'x kh', 'kh h', 'kh x', 'y j', 'y i', 'j y', 'j i', 'i j', 'i y', 'c ts', 'ts c', 'yu iu', 'iu yu', 'ya ia', 'ia ya']
    result = []
    result.append(value1)
    for liter in liters:
        replace_liters = liter.split()
        value2 = value1.replace(replace_liters[0], replace_liters[1])
        if value2 != value1:
            result.append(value2)
    return result

def search_items(keyword, data):
    result = set()  # Используем множество для исключения дубликатов

    # Приведение к нижнему регистру
    normalized_keyword = keyword.lower()

    # Транслитерация в обе стороны
    cyrillic_variants = latin_change_liters(normalized_keyword)
    latin_variants = transliterate_cyrillic_to_latin(normalized_keyword)

    # Поиск по всей строке целиком
    for item in data:
        for element in item:
            normalized_element = str(element).lower()

            # Поиск по кириллице
            if normalized_keyword in normalized_element:
                result.add(item)
                break

            # Поиск по латинице
            if any(variant in normalized_element for variant in latin_variants):
                result.add(item)
                break

            # Поиск по расшифровке английских слов
            english_variants = latin_change_liters(normalized_keyword)
            if any(variant in normalized_element for variant in english_variants):
                result.add(item)
                break

    # Поиск по каждому слову отдельно
    keywords = re.findall(r'\w+', normalized_keyword)
    for item in data:
        for element in item:
            normalized_element = str(element).lower()

            for word in keywords:
                # Поиск по кириллице
                if word in normalized_element:
                    result.add(item)
                    break

                # Поиск по латинице
                if any(variant in normalized_element for variant in latin_variants):
                    result.add(item)
                    break

                # Поиск по расшифровке английских слов
                english_variants = latin_change_liters(word)
                if any(variant in normalized_element for variant in english_variants):
                    result.add(item)
                    break

    return list(result)  # Преобразуем множество обратно в список

# ======================================================================================================

# отправка заявок по отчетам
def sendrep(message, tasks):
    tasksl = listgen(tasks, [0, 1, 3, 4, 6], 1)
    for task in tasksl:
        bot.send_message(
            message.chat.id,
            task,
            reply_markup=buttons.buttonsinline([['Показать подробности', 'tasklist '+task[0]],['ЛОКАЦИЯ', 'location '+task[0]]])
        )
    return
# Отчет в экселе


def sendrepfile(message, tasks):
    processing = bot.send_sticker(message.chat.id, "CAACAgIAAxkBAAEJL8dkedQ1ckrfN8fniwY7yUc-YNaW_AACIAAD9wLID1KiROfjtgxPLwQ", reply_markup=buttons.clearbuttons())
    rep = []
    rep.append(['№', 'ИНН', 'Контрагент', 'Заявка', 'Зарегистрирована', 'ДН', 'Менеджер', 'Выполнена', 'ДН', 'мастер'])

    wb = openpyxl.Workbook()
    ws = wb.active
    masterr = []
    top10conts = []
    taskquantity = 0

    for task in tasks:
        manager_record = db.get_record_by_id('Users', task[2])
        manager = f"{manager_record[2]} {manager_record[1]}" if manager_record is not None else ""

        master_record = db.get_record_by_id('Users', task[6])
        master = f"{master_record[2]} {master_record[1]}" if master_record is not None else ""

        contr_record = db.get_record_by_id('Contragents', task[3])
        contr = contr_record[1] if contr_record is not None else ""

        reg_date = datetime.strptime(task[1], '%d.%m.%Y %H:%M')
        reg_weekday = day_translation[reg_date.strftime('%A')]
        done_date = datetime.strptime(task[7], '%d.%m.%Y %H:%M')
        done_weekday = day_translation[done_date.strftime('%A')]

        line1 = [task[0], task[3], contr, task[4], task[1], reg_weekday, manager, task[7], done_weekday, master]
        rep.append(line1)

        if len(masterr) > 0:
            finded = 0
            for mast in masterr:
                if mast[0] == master:
                    mast[1] = mast[1] + 1
                    finded = 1
            if finded == 0:
                masterr.append([master, 1])
        else:
            masterr.append([master, 1])

        if task[3] != 303128034 and task[3] != 301263843:
            if len(top10conts) > 0:
                finded = 0
                for client in top10conts:
                    if client[0] == contr:
                        client[1] = client[1] + 1
                        finded = 1
                if finded == 0:
                    top10conts.append([contr, 1])
            else:
                top10conts.append([contr, 1])
        taskquantity = taskquantity + 1

    sorted_masterr = sorted(masterr, key=lambda x: x[1], reverse=True)
    sorted_top10conts = sorted(top10conts, key=lambda x: x[1], reverse=True)
    masterrating = '╔═════════════════╗\n  РЕЙТИНГ ПО МАСТЕРАМ:\n╚═════════════════╝\n'
    for element in sorted_masterr:
        masterrating = masterrating + '\n' + element[0] + ' - ' + str(element[1])
        pr = element[1]/(taskquantity/100)
        bar = ''
        k = 0
        l = 0
        while k < int(pr/5):
            bar = bar + '■'
            k = k + 1
        while l < 20 - int(pr/5):
            bar = bar + '□'
            l = l + 1
        masterrating = masterrating + '\n' + bar + ' ' + str(round(pr, 2)) + '%\n'
    masterrating1 = '╔════════════╗\n   Топ 10 клиентов:\n╚════════════╝\n'
    i = 0
    if len(sorted_top10conts) >= 10:
        while i < 10:
            masterrating1 = masterrating1 + '\n' + sorted_top10conts[i][0] + ' - ' + str(sorted_top10conts[i][1])
            i = i + 1
    else:
        for elem in sorted_top10conts:
            masterrating1 = masterrating1 + '\n' + elem[0] + ' - ' + str(elem[1])

    for row in rep:
        ws.append(row)

    # Установка ширины ячеек
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            cell_value = str(cell.value)
            if len(cell_value) > max_length:
                max_length = len(cell_value)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Ширина столбцов
    ws.column_dimensions['C'].width = 41.22
    ws.column_dimensions['D'].width = 41.22

    # Включение переноса слов
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Выравнивание строк
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='center', horizontal='left')

    # Выравнивание шапки
    for cell in ws[1]:
        cell.alignment = Alignment(vertical='center', horizontal='center')
        cell.font = openpyxl.styles.Font(bold=True)

    # Установка границ
    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # for row in ws.iter_rows(min_row=2):
    #     reg_date = datetime.strptime(row[4].value, '%d.%m.%Y %H:%M')  # Значение в колонке E
    #     reg_weekday = reg_date.strftime('%A')
    #     row[5].value = reg_weekday
    #     done_date = datetime.strptime(row[7].value, '%d.%m.%Y %H:%M')  # Значение в колонке H
    #     done_weekday = done_date.strftime('%A')
    #     row[8].value = done_weekday

    # Подсветка строк
    for row in ws.iter_rows(min_row=2):
        date1 = datetime.strptime(row[4].value, '%d.%m.%Y %H:%M')  # Значение в колонке E
        date2 = datetime.strptime(row[7].value, '%d.%m.%Y %H:%M')  # Значение в колонке G
        diff_hours = (date2 - date1).total_seconds() / 3600  # Разница в часах
        print(diff_hours)
        if diff_hours < 24:
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(fgColor="C4FFC4", fill_type="solid")  # Светло зеленый
        elif 24 <= diff_hours < 72:
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(fgColor="FFFFCC", fill_type="solid")  # Светло желтый
        else:
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(fgColor="FFD3DB", fill_type="solid")  # Розовый


    filename = 'rep.xlsx'
    wb.save(filename)
    bot.send_document(message.chat.id, open(filename, 'rb'))
    bot.send_message(
        message.chat.id,
        masterrating
    )
    bot.send_message(
        message.chat.id,
        masterrating1
    )
    os.remove(filename)
    bot.delete_message(message.chat.id, processing.message_id)



# Удаление сообщений о новой заявке
def deletentm(taskid):
    messages = db.select_table_with_filters('NewTasksMessages', {'taskid': taskid})
    for mes in messages:
        try:
            mesdel(mes[2], mes[3])
            db.delete_record('NewTasksMessages', 'id', mes[0])
        except Exception as e:
            pass
