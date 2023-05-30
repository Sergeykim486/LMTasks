import openpyxl, os, config, telebot, functions, buttons, logging, time, datetime, asyncio, threading
from telebot import TeleBot, types
from db import Database
from datetime import datetime
from openpyxl.styles import Alignment

# логи
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
# Глобальная переменная, база и создание объекта бот
ActiveUser = {}
sendedmessages = []
dbname = os.path.dirname(os.path.abspath(__file__)) + '/Database/' + 'lmtasksbase.db'
db = Database(dbname)
bot = telebot.TeleBot(config.TOKEN)
continue_polling = True
sch = 0
# Добавление в список пользователей клиентского бота
if db.get_record_by_id('Users', 0) == None:
    db.insert_record(
        'Users',
        [
            0,
            'Client BOT',
            '...',
            '...'
        ]
    )
# Добавление новой таблицы с локациями и голонки локации для заявки
db.add_column_to_table("Tasks", "location", "INTEGER")
cols = [
    "id INTEGER PRIMARY KEY",
    "inn INTEGER",
    "name TEXT",
    "lat TEXT",
    "lon TEXT"
]
db.create_table("Locations", cols)
try:
    if db.get_record_by_id('Statuses', 5) is None and db.get_record_by_id('Statuses', 6) is None and db.get_record_by_id('Statuses', 7) is None:
        db.insert_record('Statuses', [5, 'Зарегистрирована'])
        db.insert_record('Statuses', [6, 'Прринята мастером'])
        db.insert_record('Statuses', [7, 'Выполнено'])
except Exception as e:
    logging.error(e)
    pass
try:
    if db.get_record_by_id('locations', 999) is None:
        db.insert_record('locations', [999, 0, 'Техника в офисе', 41.28921489333344, 69.31288111459628])
except Exception as e:
    logging.error(e)
    pass
# Проверка расписания
async def job():
    await schedule_message()
async def schedule_message():
    while True:
        try:
            Tasks = db.select_table_with_filters('Tasks', {'status': 0})
            if len(Tasks) > 0:
                for line in Tasks:
                    db.update_records('Tasks', ['status'], [1], 'id', line[0])
                    tid = line[0]
                    sendtoall(functions.curtask(tid), buttons.buttonsinline([['👍 Принять', 'confirm ' + str(tid)], ['📎 Назначить', 'set ' + str(tid)]]), 0)
        except Exception as e:
            logging.error(e)
            pass
        try:
            revs = db.select_table_with_filters('rev', {'status': 0})
            if len(revs) > 0:
                for line in revs:
                    db.update_records('rev', ['status'], [1], 'id', line[0])
                    mes = 'Поступил отзыв/оценка от клиента\n'
                    mes = mes + '\nКЛИЕНТ - ' + str(db.get_record_by_id('Clients', line[2])[2])
                    mes = mes + '\n\nОТЗЫВ:\n' + str(line[3])
                    mes = mes + '\n\nот ' + str(line[1])
                    sendtoall(mes, '', 0)
        except Exception as e:
            logging.error(e)
            pass
        now = datetime.now()
        # if now.hour == 8 and now.minute == 0:
        #     await daylyreport.morning()
        if now.hour == 20 and now.minute == 0:
            await daylyreport.evening()
        daterep = str(datetime.now().strftime("%d.%m.%Y"))
        locations = []
        addedlocs = db.select_table_with_filters('Tasks', {'status': 1})
        conflocs = db.select_table_with_filters('Tasks', {'status': 2})
        donet = db.select_table_with_filters('Tasks', {'status': 3}, ['done'], [daterep+' 00:00'], [daterep+' 23:59'])
        canceled = db.select_table_with_filters('Tasks', {'status': 4}, ['canceled'], [daterep+' 00:00'], [daterep+' 23:59'])
        try:
            for task in addedlocs:
                company = db.get_record_by_id('Contragents', task[3])[1]
                status = db.get_record_by_id('Statuses', task[11])[1]
                name = '№ ' + str(task[0]) + '\n|=============================|\n' + str(company)
                description = status + '\n | \n' + task[4]
                location = db.get_record_by_id('Locations', task[12])
                if task[12] is not None and location is not None:
                    lat = location[3]
                    lon = location[4]
                else:
                    lat = 41.28921489333344
                    lon = 69.31288111459628
                locations.append([name, description, lat, lon, task[11]])
        except Exception as e:
            logging.info(e)
            pass
        try:
            for task in conflocs:
                company = db.get_record_by_id('Contragents', task[3])[1]
                status = db.get_record_by_id('Statuses', task[11])[1]
                user = db.get_record_by_id('Users', task[6])
                master = str(user[2]) + ' ' + str(user[1])
                name = '№ ' + str(task[0]) + '\n|=============================|\n' + str(company)
                description = status + ' - ' + master + '\n | \n' + task[4]
                if task[12] is None:
                    lat = 41.28921489333344
                    lon = 69.31288111459628
                else:
                    lat = db.get_record_by_id('Locations', task[12])[3]
                    lon = db.get_record_by_id('Locations', task[12])[4]
                locations.append([name, description, lat, lon, task[11]])
        except Exception as e:
            logging.info(e)
            pass
        try:
            for task in donet:
                company = db.get_record_by_id('Contragents', task[3])[1]
                status = db.get_record_by_id('Statuses', task[11])[1]
                user = db.get_record_by_id('Users', task[6])
                master = str(user[2]) + ' ' + str(user[1])
                name = '№ ' + str(task[0]) + '\n|=============================|\n' + str(company)
                description = status + ' - ' + master + '\n | \n' + task[4]
                if task[12] is None:
                    lat = 41.28921489333344
                    lon = 69.31288111459628
                else:
                    lat = db.get_record_by_id('Locations', task[12])[3]
                    lon = db.get_record_by_id('Locations', task[12])[4]
                locations.append([name, description, lat, lon, task[11]])
        except Exception as e:
            logging.info(e)
            pass
        try:
            for task in canceled:
                company = db.get_record_by_id('Contragents', task[3])[1]
                status = db.get_record_by_id('Statuses', task[11])[1]
                name = '№ ' + str(task[0]) + '\n|=============================|\n' + str(company)
                description = status + '\n | \n' + task[4]
                if task[12] is None:
                    lat = 41.28921489333344
                    lon = 69.31288111459628
                else:
                    lat = db.get_record_by_id('Locations', task[12])[3]
                    lon = db.get_record_by_id('Locations', task[12])[4]
                locations.append([name, description, lat, lon, task[11]])
        except Exception as e:
            logging.info(e)
            pass
        if len(locations) > 0:
            functions.mmapgen(locations)
            functions.mapgen(locations)
        await asyncio.sleep(15)
async def main():
    await job()
# отправка сообщения всем пользователям
def sendtoall(message, markdown, exeptions, nt = 0, notific = False):
    global sendedmessages
    users = db.select_table('Users')
    for user in users:
        logging.info(f'sended message to user {user[2]} {user[1]}')
        try:
            if user[0] != exeptions:
                mes = bot.send_message(
                    user[0],
                    message,
                    reply_markup=markdown,
                    disable_notification=notific
                )
                if nt == 1:
                    sendedmessages.append([[user[0]], [mes.message_id]])
        except Exception as e:
            logging.error(e)
            pass
    return
# Список локаций контрагента
def sendlocations(inn, message):
    locs = db.select_table_with_filters('Locations', {'inn': inn})
    if len(locs) > 0:
        for location in locs:
            loc = types.Location(location[4], location[3])
            bot.send_location(message.chat.id, loc.latitude, loc.longitude)
            bot.send_message(
                message.chat.id,
                str(location[2]),
                reply_markup=buttons.buttonsinline([['Изменить', 'location '+str(location[0])]])
            )
    else:
        return
# отправка заявок по отчетам
def sendrep(message, tasks):
    tasksl = functions.listgen(tasks, [0, 1, 3, 4, 6], 1)
    for task in tasksl:
        bot.send_message(
            message.chat.id,
            task,
            reply_markup=buttons.buttonsinline([['Показать подробности', 'tasklist '+task[0]]])
        )
    return
# Отправка отчета в виде таблицы
# def sendrepfile(message, tasks):
#     processing = bot.send_message(message.chat.id, '⏳')
#     rep = []
#     # шапка
#     rep.append(['№', 'ИНН', 'Контрагент', 'Заявка', 'Зарегистрирована', 'Менеджер', 'Выполнена', 'мастер'])

#     # Объединение ячеек в шапке
#     wb = openpyxl.Workbook()
#     ws = wb.active

#     for task in tasks:
#         manager = str(db.get_record_by_id('Users', task[2])[2]) + ' ' + str(db.get_record_by_id('Users', task[2])[1])
#         master = str(db.get_record_by_id('Users', task[6])[2]) + ' ' + str(db.get_record_by_id('Users', task[6])[1])
#         contr = (str(db.get_record_by_id('Contragents', task[3])[1]) if db.get_record_by_id('Contragents', task[3]) is not None else '')
#         line1 = [task[0], task[3], contr, task[4], task[1], manager, task[7], master]
#         rep.append(line1)

#     for row in rep:
#         ws.append(row)

#     # Установка ширины ячеек
#     for column_cells in ws.columns:
#         max_length = 0
#         column = column_cells[0].column_letter
#         for cell in column_cells:
#             cell_value = str(cell.value)
#             if len(cell_value) > max_length:
#                 max_length = len(cell_value)
#         adjusted_width = (max_length + 2) * 1.2
#         ws.column_dimensions[column].width = adjusted_width

#     # Ширина столбцов C и D
#     ws.column_dimensions['C'].width = 41.22
#     ws.column_dimensions['D'].width = 41.22

#     # Включение переноса слов во всех ячейках
#     for row in ws.iter_rows():
#         for cell in row:
#             cell.alignment = Alignment(wrap_text=True)

#     # Выравнивание строк
#     for row in ws.iter_rows(min_row=2):
#         for cell in row:
#             cell.alignment = Alignment(vertical='center', horizontal='left')

#     # Выравнивание шапки
#     for cell in ws[1]:
#         cell.alignment = Alignment(vertical='center', horizontal='center')
#         cell.font = openpyxl.styles.Font(bold=True)

#     # Установка границ
#     thin_border = openpyxl.styles.Border(
#         left=openpyxl.styles.Side(style='thin'),
#         right=openpyxl.styles.Side(style='thin'),
#         top=openpyxl.styles.Side(style='thin'),
#         bottom=openpyxl.styles.Side(style='thin')
#     )
#     for row in ws.iter_rows():
#         for cell in row:
#             cell.border = thin_border

#     file_path = os.path.join(os.getcwd(), 'data.xlsx')
#     wb.save(file_path)
#     bot.delete_message(chat_id=message.chat.id, message_id=processing.message_id)
#     bot.send_document(message.chat.id, open(file_path, 'rb'))
#     os.remove(file_path)


def sendrepfile(message, tasks):
    processing = bot.send_message(message.chat.id, '⏳')
    rep = []
    # шапка
    rep.append(['№', 'ИНН', 'Контрагент', 'Заявка', 'Зарегистрирована', 'Менеджер', 'Выполнена', 'мастер'])

    # Объединение ячеек в шапке
    wb = openpyxl.Workbook()
    ws = wb.active

    for task in tasks:
        manager = str(db.get_record_by_id('Users', task[2])[2]) + ' ' + str(db.get_record_by_id('Users', task[2])[1])
        master = str(db.get_record_by_id('Users', task[6])[2]) + ' ' + str(db.get_record_by_id('Users', task[6])[1])
        contr = (str(db.get_record_by_id('Contragents', task[3])[1]) if db.get_record_by_id('Contragents', task[3]) is not None else '')
        line1 = [task[0], task[3], contr, task[4], task[1], manager, task[7], master]
        rep.append(line1)

    for row in rep:
        ws.append(row)

    # Установка ширины ячеек
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            cell_value = str(cell.value)
            if len(cell_value) > max_length:
                max_length = len(cell_value)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Ширина столбцов C и D
    ws.column_dimensions['C'].width = 41.22
    ws.column_dimensions['D'].width = 41.22

    # Включение переноса слов во всех ячейках
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Выравнивание строк
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='center', horizontal='left')

    # Выравнивание шапки
    for cell in ws[1]:
        cell.alignment = Alignment(vertical='center', horizontal='center')
        cell.font = openpyxl.styles.Font(bold=True)

    # Установка границ
    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Подсветка строк
    for row in ws.iter_rows(min_row=2):
        date1 = datetime.strptime(row[4].value, '%d.%m.%Y %H:%M')  # Значение в колонке E
        date2 = datetime.strptime(row[6].value, '%d.%m.%Y %H:%M')  # Значение в колонке G
        diff_hours = (date2 - date1).total_seconds() / 3600  # Разница в часах
        if diff_hours < 24:
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(fgColor="C4FFC4", fill_type="solid")  # Светло зеленый
        elif 24 <= diff_hours < 72:
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(fgColor="FFFFCC", fill_type="solid")  # Светло желтый
        else:
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(fgColor="FFD3DB", fill_type="solid")  # Розовый

    file_path = os.path.join(os.getcwd(), 'data.xlsx')
    wb.save(file_path)
    bot.delete_message(chat_id=message.chat.id, message_id=processing.message_id)
    bot.send_document(message.chat.id, open(file_path, 'rb'))
    os.remove(file_path)


# Дневной отчет для рассылки по расписанию
class daylyreport:
    # Рассылка текущих хвостов спредыдущих дней
    async def morning():
        logging.info('план отправлен.')
        confirmedtasks = functions.listgen(db.select_table_with_filters('Tasks', {'status': 2}), [0, 1, 3, 4, 6], 1)
        addedtasks = functions.listgen(db.select_table_with_filters('Tasks', {'status': 1}), [0, 1, 3, 4, 6], 1)
        if len(confirmedtasks) == 0 and len(addedtasks) == 0:
            sendtoall('Всем доброе утро!\nНа сегодня нет переходящих заявок.', '', 0)
        else:
            sendtoall('Всем доброе утро!\nСо вчерашнего дня на сегодня переходят следующие заявки:', '', 0)
        if len(confirmedtasks) != 0:
            sendtoall('ЗАЯВКИ У МАСТЕРОВ:\n🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻', '', 0)
            for line in confirmedtasks:
                taskid = line.split()[2]
                sendtoall(line, buttons.buttonsinline([['Показать подробности', 'tasklist '+taskid]]), 0)
        if len(addedtasks) != 0:
            sendtoall('НЕ РАСПРЕДЕЛЕННЫЕ ЗАЯВКИ:\n🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻', '', 0)
            for line in addedtasks:
                taskid = line.split()[2]
                sendtoall(line, buttons.buttonsinline([['Показать подробности', 'tasklist '+taskid]]), 0)
        sendtoall('🟥🟥🟥🟥🟥🟥🟥🟥\nСписок заявок на сегодня\n🟥🟥🟥🟥🟥🟥🟥🟥', '', 0)
    # Итоги дня
    async def evening():
        logging.info('план отправлен.')
        daten = str(datetime.now().strftime("%d.%m.%Y"))
        donetasks = functions.listgen(db.select_table_with_filters('Tasks', {'status': 3}, ['done'], [daten+' 00:00'], [daten+' 23:59']), [0, 1, 3, 4, 6], 1)
        confirmedtasks = functions.listgen(db.select_table_with_filters('Tasks', {'status': 2}), [0, 1, 3, 4, 6], 1)
        addedtasks = functions.listgen(db.select_table_with_filters('Tasks', {'status': 1}), [0, 1, 3, 4, 6], 1)
        canceledtasks = functions.listgen(db.select_table_with_filters('Tasks', {'status': 4}, ['canceled'], [daten+' 00:00'], [daten+' 23:59']), [0, 1, 3, 4, 6], 1)
        if len(confirmedtasks) != 0 and len(addedtasks) != 0:
            sendtoall('ИТОГИ ДНЯ:\nНа завтра остаются следующие заявки:', '', 0)
        if len(donetasks) != 0:
            sendtoall('Выполненные заявки\n🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻', '', 0)
            for line in donetasks:
                taskid = line.split()[2]
                sendtoall(line, buttons.buttonsinline([['Показать подробности', 'tasklist '+taskid]]), 0)
        if len(confirmedtasks) != 0:
            sendtoall('Заявки у мастеров\n🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻', '', 0)
            for line in confirmedtasks:
                taskid = line.split()[2]
                sendtoall(line, buttons.buttonsinline([['Показать подробности', 'tasklist '+taskid]]), 0)
        if len(addedtasks) != 0:
            sendtoall('Не принятые заявки\n🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻', '', 0)
            for line in addedtasks:
                taskid = line.split()[2]
                sendtoall(line, buttons.buttonsinline([['Показать подробности', 'tasklist '+taskid]]), 0)
        if len(canceledtasks) != 0:
            sendtoall('Отмененные\n🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻', '', 0)
            for line in canceledtasks:
                taskid = line.split()[2]
                sendtoall(line, buttons.buttonsinline([['Показать подробности', 'tasklist '+taskid]]), 0)
        reports = '\nВыполнено - ' + str(len(donetasks)) + '\nНе распределенных - ' + str(len(addedtasks)) + '\nВ работе у мастеров - ' + str(len(confirmedtasks)) + '\nОтменено - ' + str(len(canceledtasks))
        if len(donetasks) == 0:
            reports = reports + '\n\nВыполненных заявок нет.'
        else:
            reports = reports + '\n\nКоличество заявок выполненных мастерами:\n'
            users = db.select_table('Users')
            usersrep = []
            for i in users:
                tasks = len(db.select_table_with_filters('Tasks', {'master': i[0], 'status': 3}, ['done'], [daten+' 00:00'], [daten+' 23:59']))
                usersrep.append([i[2] + ' ' + i[1], tasks])
            sorted_usersrep = sorted(usersrep, key=lambda x: x[1], reverse=True)
            for j in sorted_usersrep:
                if j[1] != 0:
                    reports = reports + '\n' + j[0] + ' - ' + str(j[1])
        sendtoall('ИТОГИ ДНЯ\n🔺🔺🔺🔺🔺🔺🔺🔺🔺🔺🔺🔺' + reports, '', 0)

@bot.message_handler(commands=['start'])
# проверка пользователя при первом запуске
def check_user_id(message):
    user_id = message.from_user.id
    global ActiveUser, continue_polling
    try:
        username = db.get_record_by_id('Users', user_id)[2] + ' ' + db.get_record_by_id('Users', user_id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
    except Exception as e:
        logging.error(e)
        pass
    continue_polling = True
    ActiveUser[user_id] = {'id': user_id}
    # finduser = db.search_record("Users", "id", user_id)
    user = db.get_record_by_id('Users', user_id)
    if user is None:
        bot.send_message(
            user_id,
            'Вам нужно пройти регистрацию',
            reply_markup=buttons.Buttons(['🔑 Регистрация'])
        )
        bot.register_next_step_handler(message, Reg.reg1)

    else:
        bot.send_message(
            user_id,
            'Выберите операцию.',
            reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
        )
        bot.register_next_step_handler(message, MainMenu.Main2)

@bot.message_handler(func=lambda message: True)
# проверка пользователя при первом запуске
def check_user_id(message):
    user_id = message.from_user.id
    global ActiveUser, continue_polling
    try:
        username = db.get_record_by_id('Users', user_id)[2] + ' ' + db.get_record_by_id('Users', user_id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
    except Exception as e:
        logging.error(e)
        pass
    continue_polling = True
    ActiveUser[user_id] = {'id': user_id}
    # finduser = db.search_record("Users", "id", user_id)
    user = db.get_record_by_id('Users', user_id)
    if user is None:
        bot.send_message(
            user_id,
            'Вам нужно пройти регистрацию',
            reply_markup=buttons.Buttons(['🔑 Регистрация'])
        )
        bot.register_next_step_handler(message, Reg.reg1)

# Регистрация нового пользователя
class Reg:
    # Запрос имени у пользователя
    def reg1(message):
        try:
            username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
            logging.info(f'{username} Отправил запрос - {message.text}')
        except Exception as e:
            logging.error(e)
            pass
        if message.text == '🔑 Регистрация':
            bot.send_message(
                message.chat.id,
                'Как Вас зовут (укажите имя)',
            reply_markup=buttons.clearbuttons()
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, Reg.reg2)
        else:
            bot.send_message(
                message.chat.id,
                'Пожалуйста зарегистрируйтесь.',
                reply_markup=buttons.Buttons(['🔑 Регистрация'])
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, Reg.reg1)
    # Сохранение имени и запрос фамилии
    def reg2(message):
        try:
            username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
            logging.info(f'{username} Отправил запрос - {message.text}')
        except Exception as e:
            logging.error(e)
            pass
        global ActiveUser
        ActiveUser[message.chat.id]['FirstName'] = message.text
        bot.send_message(
            message.chat.id,
            'Укажите Вашу фамилию.',
            reply_markup=buttons.clearbuttons()
        )
        bot.register_next_step_handler(message, Reg.reg3)
    # Сохранение фамилии и запрос номера телефона
    def reg3(message):
        try:
            username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
            logging.info(f'{username} Отправил запрос - {message.text}')
        except Exception as e:
            logging.error(e)
            pass
        global ActiveUser
        ActiveUser[message.chat.id]['LastName'] = message.text
        bot.send_message(
            message.chat.id,
            'Введите Ваш номер телефона в формате (+998 00 000 0000).',
            reply_markup=buttons.clearbuttons()
        )
        bot.register_next_step_handler(message, Reg.reg4)
    # Сохранение телефона и запрос на подтверждение сохраненных данных
    def reg4(message):
        try:
            username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
            logging.info(f'{username} Отправил запрос - {message.text}')
        except Exception as e:
            logging.error(e)
            pass
        global ActiveUser
        ActiveUser[message.chat.id]['PhoneNumber'] = message.text
        bot.send_message(
            message.chat.id,
            functions.conftext(message, ActiveUser),
            reply_markup=buttons.Buttons(['✅ Да', '⛔️ Нет'])
        )
        bot.register_next_step_handler(message, Reg.reg5)
    # Проверка подтверждения данных
    def reg5(message):
        try:
            username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
            logging.info(f'{username} Отправил запрос - {message.text}')
        except Exception as e:
            logging.error(e)
            pass
        global ActiveUser
        if message.text == '✅ Да':
            valuedict = [
                ActiveUser[message.chat.id]['id'],
                ActiveUser[message.chat.id]['FirstName'],
                ActiveUser[message.chat.id]['LastName'],
                ActiveUser[message.chat.id]['PhoneNumber']
            ]
            db.insert_record("Users", valuedict)
            bot.send_message(
                message.chat.id,
                'Поздравляем Вы успешно зарегистрировались!',
                reply_markup=buttons.Buttons(['🏠 Главное меню'])
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, MainMenu.Main1)
        elif message.text == '⛔️ Нет':
            bot.send_message(
                message.chat.id,
                'Пройдите регистрацию повторно.',
                reply_markup=buttons.Buttons(['🔑 Регистрация'])
            )
            bot.register_next_step_handler(message, Reg.reg1)
        else:
            bot.send_message(
                message.chat.id,
                'Вы не подтвердили информацию!\n' + functions.conftext(message, ActiveUser),
                reply_markup=buttons.Buttons(['✅ Да', '⛔️ Нет'])
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, Reg.reg5)
# Главное меню и обработка кнопок главного меню
class MainMenu:
    # Главное меню
    def Main1(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        if message.text == '🏠 Главное меню' or message.text == 'Вернуться':
            ActiveUser[message.chat.id]['sentmes'] = bot.send_message(
                message.chat.id,
                'Выберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, MainMenu.Main2)
    # Реакия на кнопки главного меню
    def Main2(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser, continue_polling
        if message.text == '📝 Новая заявка':
            ActiveUser[message.chat.id]['nt'] = 1
            ActiveUser[message.chat.id]['sentmes'] = bot.send_message(
                message.chat.id,
                'Введите ИНН, ПИНФЛ или серию пвсспоррта клиента.\nТак же Вы можете попытаться поискать контрагента по наименованию или его части\nНапримар:\nmonohrom\nВыдаст все компании из базы бота у которых в названии есть monohrom',
                reply_markup=buttons.Buttons(['🚫 Отмена'])
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, NewTask.nt1)
        elif message.text == '🔃 Обновить список заявок':
            daterep = str(datetime.now().strftime("%d.%m.%Y"))
            report.rep(message, daterep, 0, 1, 1, 0, 0)
        elif message.text == '🖨️ Обновить список техники':
            daterep = str(datetime.now().strftime("%d.%m.%Y"))
            report.rep(message, daterep, 0, 0, 0, 0, 0, 0, 0, 1, 1, 1)
        elif message.text == '📋 Мои заявки':
            daterep = str(datetime.now().strftime("%d.%m.%Y"))
            report.rep(message, daterep, 0, 1, 0, 1, 0, message.chat.id, 1)
        elif message.text == '/start':
            bot.send_message(
                message.chat.id,
                'Выберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.register_next_step_handler(message, MainMenu.Main2)
        elif message.text == '📢 Написать всем':
            bot.send_message(
                message.chat.id,
                'Напишите Ваше сообщение и оно будет разослано всем.\nчтобы вернуться в главное меню нажмите [Главное меню]',
                reply_markup=buttons.Buttons(['🏠 Главное меню'])
            )
            if message.message_id is not None:
                bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)

            bot.register_next_step_handler(message, allchats.chat1)
        elif message.text == '📈 Отчеты':
            bot.send_message(
                message.chat.id,
                'Выберите какой отчет Вам нужен\nПоказать все не выполненные заявки, или показать итоги дня.',
                reply_markup=buttons.Buttons(['📋 Заявки у мастеров', '🖨️ Техника у мастеров', '📊 Итоги дня', '📆 За период', '🚫 Отмена'])
            )
            if message.message_id is not None:
                bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, report.reportall)
        elif message.text == '✏️ Редактировать контрагента':
            contragents = db.select_table('Contragents', ['id', 'cname'])
            ActiveUser[message.chat.id]['sentmes'] = bot.send_message(
                message.chat.id,
                'Введите ИНН клиента.\nИли вы можете поискать по названию',
                reply_markup=buttons.Buttons(['🚫 Отмена'])
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, editcont.ec1)
        elif message.text == '🗺️ Карта':
            markup = types.InlineKeyboardMarkup()
            button = types.InlineKeyboardButton(text='Открыть карту', url='http://81.200.149.148/map.html')
            markup.add(button)
            bot.send_message(
                message.chat.id,
                'Вы можете посмотреть все теущие заявки за сегодня, на карте',
                reply_markup=markup
            )
            bot.register_next_step_handler(message, MainMenu.Main2)
        elif message.text is not None:
            if message.text.isdigit() or (len(message.text.split()) > 1 and message.text.split()[1].isdigit()):
                if message.text.isdigit():
                    taskid = message.text
                elif message.text.split()[1].isdigit():
                    taskid = message.text.split()[1]
                task = db.get_record_by_id('Tasks', taskid)
                tasks = functions.listgen([task], [0, 1, 3, 4, 6], 1)
                if task is not None:
                    bot.send_message(
                        message.chat.id,
                        tasks[0],
                        reply_markup=buttons.buttonsinline([['Показать подробности', 'tasklist '+taskid]])
                    )
            ActiveUser[message.chat.id]['sentmes'] = bot.send_message(
                message.chat.id,
                'Выберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.register_next_step_handler(message, MainMenu.Main2)
        else:
            bot.register_next_step_handler(message, MainMenu.Main2)
# Редактирование контрагента
class editcont():
    # Поиск контрагента по ИНН и генераия основной формы def editcontragent(message)
    def ec1(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        if message.text == '🚫 Отмена':
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.delete_message(chat_id=ActiveUser[message.chat.id]['sentmes'].chat.id, message_id=ActiveUser[message.chat.id]['sentmes'].message_id)
            bot.send_message(
                message.chat.id,
                'Выберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.register_next_step_handler(message, MainMenu.Main2)
        elif message.text.split()[0].isdigit():
            inn = message.text.split()[0]
            ActiveUser[message.chat.id]['inn'] = inn
            client = db.get_record_by_id('Contragents', ActiveUser[message.chat.id]['inn'])
            ActiveUser[message.chat.id]['contold'] = client
            ActiveUser[message.chat.id]['contnew'] = []
            for line in client:
                if line is None:
                    ActiveUser[message.chat.id]['contnew'].append(None)
                else:
                    ActiveUser[message.chat.id]['contnew'].append(line)
            if client is not None:
                editcontragent(message)
                bot.register_next_step_handler(message, editcont.ec2)
            else:
                bot.send_message(
                    message.chat.id,
                    'Контрагент не найден.',
                    reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
                )
                bot.register_next_step_handler(message, MainMenu.Main2)
        else:
            processing = bot.send_message(message.chat.id, '⏳')
            contrs = db.select_table('Contragents')
            res = functions.search_items(message.text, contrs)
            contbuttons = []
            contbuttons.append('🚫 Отмена')
            if len(res) > 0:
                for i in res:
                    line = str(i[0]) + ' ' + str(i[1])
                    if len(contbuttons) < 20:
                        contbuttons.append(line)
                bot.delete_message(chat_id=message.chat.id, message_id=processing.message_id)
                try:
                    bot.send_message(
                        message.chat.id,
                        'Если нужный клиент не вышел в списке, попробуйте перефразировать и ввести снова.\nВыберите контрагента из списка, или введите его ИНН, ПИНФЛ, серию пасспорта или повторите поиск.',
                        reply_markup=buttons.Buttons(contbuttons, 1)
                    )
                except Exception as e:
                    logging.error(e)
                    pass
            else:
                bot.delete_message(chat_id=message.chat.id, message_id=processing.message_id)
                bot.send_message(
                    message.chat.id,
                    'Контрагент не найден.',
                    reply_markup=buttons.Buttons(contbuttons, 1)
                )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, editcont.ec1)
    # Реакция на нажатие кнопок в меню редактирования
    def ec2(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        if message.text == '💾 Сохранить':
            print(ActiveUser[message.chat.id]['contnew'])
            if ActiveUser[message.chat.id]['contnew'][0] != ActiveUser[message.chat.id]['contold'][0]:
                tasks = db.select_table_with_filters('Tasks', {'contragent': ActiveUser[message.chat.id]['contold'][0]})
                for task in tasks:
                    db.update_records(
                        'Tasks',
                        ['contragent'],
                        [ActiveUser[message.chat.id]['contnew'][0]],
                        'id',
                        task[0]
                    )
                locations = db.select_table_with_filters('Locations', {'inn': ActiveUser[message.chat.id]['contold'][0]})
                for location in locations:
                    db.update_records(
                        'Locations',
                        ['inn'],
                        [ActiveUser[message.chat.id]['contnew'][0]],
                        'id',
                        location[0]
                    )
            db.update_records(
                'Contragents',
                [
                    'id',
                    'cname',
                    'cadr',
                    'cperson',
                    'cphone',
                    'ds',
                    'contract'
                ],
                ActiveUser[message.chat.id]['contnew'],
                'id',
                ActiveUser[message.chat.id]['contold'][0]
            )
            bot.send_message(
                message.chat.id,
                'Данные изменены.\nВыберите действие',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.register_next_step_handler(message, MainMenu.Main2)
        elif message.text == '🚫 Отмена':
            bot.send_message(
                message.chat.id,
                'Выберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.register_next_step_handler(message, MainMenu.Main2)
        elif message.text == '🏷️ ТИП':
            ActiveUser[message.chat.id]['sentmes'] = bot.send_message(
                message.chat.id,
                f'Введите новое значение ({message.text})',
                reply_markup=buttons.Buttons(['Разовый', 'Долгосрочный', 'Физ. лицо'])
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, editcont.TYPE)
        elif message.text == '🛣️ АДРЕС':
            ActiveUser[message.chat.id]['sentmes'] = bot.send_message(
                message.chat.id,
                'Введите новый адрес или отправьте локацию.',
                reply_markup=buttons.clearbuttons()
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, CADR1)
        elif message.text == '📍 ЛОКАЦИИ':
            # ИЗМЕНЕНИЕ ЛОКАЦИЙ КОНТРАГЕНТА
            locations = db.select_table_with_filters('Locations', {'inn': ActiveUser[message.chat.id]['inn']})
            buttonsloc = []
            buttonsloc.append('🆕 Добавить')
            try:
                for location in locations:
                    buttonsloc.append(str(location[0]) + ' ' + str(location[2]))
            except Exception as e:
                logging.error(e)
                time.sleep(5)
            buttonsloc.append('↩️ Назад')
            if len(locations) > 2:
                bot.send_message(
                    message.chat.id,
                    'Это все локации выберите ту которую хотите изменить.',
                    reply_markup=buttons.Buttons(buttonsloc, 2)
                )
                bot.register_next_step_handler(message, editcont.locations1)
            else:
                bot.send_message(
                    message.chat.id,
                    'Для выбранного контрагента не указаны локации.',
                    reply_markup=buttons.Buttons(buttonsloc, 2)
                )
                bot.register_next_step_handler(message, editcont.locations1)
        else:
            ActiveUser[message.chat.id]['sentmes'] = bot.send_message(
                message.chat.id,
                f'Введите новое значение ({message.text})',
                reply_markup=buttons.clearbuttons()
            )
            if message.text == '🆔 ИНН':
                bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
                bot.register_next_step_handler(message, editcont.INN)
            elif message.text == '🏢 НАИМЕНОВАНИЕ':
                bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
                bot.register_next_step_handler(message, editcont.CNAME)
            elif message.text == '🙋‍♂️ КОНТАКТНОЕ ЛИЦО':
                bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
                bot.register_next_step_handler(message, editcont.CPERSON)
            elif message.text == '📞 ТЕЛЕФОН':
                bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
                bot.register_next_step_handler(message, editcont.CPHONE)
            elif message.text == '📄 ДОГОВОР':
                bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
                bot.register_next_step_handler(message, editcont.CCONTRACT)
    # ИНН
    def INN(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        if message.text.isdigit():
            ActiveUser[message.chat.id]['contnew'][0] = message.text
            editcontragent(message)
            bot.register_next_step_handler(message, editcont.ec2)
        else:
            bot.send_message(
                message.chat.id,
                'Не верно указан ИНН! \nПовторите попытку',
                reply_markup=buttons.clearbuttons()
            )
            bot.register_next_step_handler(message, editcont.INN)
    # Наименование организаии
    def CNAME(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        ActiveUser[message.chat.id]['contnew'][1] = message.text
        editcontragent(message)
        bot.register_next_step_handler(message, editcont.ec2)
    # Тип договора разовый или долгосрочный
    def TYPE(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        if message.text == 'Разовый':
            ActiveUser[message.chat.id]['contnew'][5] = 1
        elif message.text == 'Долгосрочный':
            ActiveUser[message.chat.id]['contnew'][5] = 2
        elif message.text == 'Физ. лицо':
            ActiveUser[message.chat.id]['contnew'][5] = 3
        editcontragent(message)
        bot.register_next_step_handler(message, editcont.ec2)
    # Контактное лио
    def CPERSON(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        ActiveUser[message.chat.id]['contnew'][3] = message.text
        editcontragent(message)
        bot.register_next_step_handler(message, editcont.ec2)
    # Контактный телефон
    def CPHONE(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        ActiveUser[message.chat.id]['contnew'][4] = message.text
        editcontragent(message)
        bot.register_next_step_handler(message, editcont.ec2)
    # Номер и дата договора (если долгосрочный)
    def CCONTRACT(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        ActiveUser[message.chat.id]['contnew'][6] = message.text
        editcontragent(message)
        bot.register_next_step_handler(message, editcont.ec2)
    # Меню и список локаций
    def locations1(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        if message.text == '🆕 Добавить':
            bot.send_message(
                message.chat.id,
                'Отправьте локацию',
                reply_markup=buttons.clearbuttons()
            )
            bot.register_next_step_handler(message, newlocation)
        elif message.text == '↩️ Назад':
            editcontragent(message)
            bot.register_next_step_handler(message, editcont.ec2)
        elif message.text.split()[0].isdigit() and db.get_record_by_id('Locations', message.text.split()[0]) is not None:
            location = db.get_record_by_id('Locations', message.text.split()[0])
            loc = types.Location(location[4], location[3])
            bot.send_location(message.chat.id, loc.latitude, loc.longitude)
            ActiveUser[message.chat.id]['curlocation'] = location[0]
            bot.send_message(
                message.chat.id,
                f'Выбрана локация: {location[2]}',
                reply_markup=buttons.Buttons(['Изменить локацию', 'Изменить название', '🗑️ Удалить', '🚫 Отмена'], 3)
            )
            bot.register_next_step_handler(message, editcont.locations2)
        else:
            bot.send_message(
                message.chat.id,
                'Не верная команда.'
            )
            bot.register_next_step_handler(message, editcont.locations1)
    # Редактирование выбранной локации
    def locations2(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        if message.text == 'Изменить локацию':
            bot.send_message(
                message.chat.id,
                'Отправьте новую локацию.',
                reply_markup=buttons.clearbuttons()
            )
            bot.register_next_step_handler(message, editcontlocation1)
        elif message.text == 'Изменить название':
            bot.send_message(
                message.chat.id,
                'Укажите новое название локации.',
                reply_markup=buttons.clearbuttons()
            )
            bot.register_next_step_handler(message, editcont.locations3)
        elif message.text == '🗑️ Удалить':
            locationtodelete = db.get_record_by_id('Locations', ActiveUser[message.chat.id]['curlocation'])[2]
            Contrlocation = db.get_record_by_id('Contragents', ActiveUser[message.chat.id]['inn'])[1]
            bot.send_message(
                message.chat.id,
                f'Удалить локацию {locationtodelete} у {Contrlocation}?',
                reply_markup=buttons.Buttons(['✅ Да','⛔️ Нет'])
            )
            bot.register_next_step_handler(message, editcont.locations4)
        elif message.text == '🚫 Отмена':
            locations = db.select_table_with_filters('Locations', {'inn': ActiveUser[message.chat.id]['inn']})
            buttonsloc = []
            buttonsloc.append('🆕 Добавить')
            try:
                for location in locations:
                    buttonsloc.append(str(location[0]) + ' ' + str(location[2]))
            except Exception as e:
                logging.error(e)
                time.sleep(5)
            buttonsloc.append('↩️ Назад')
            if len(locations) > 2:
                bot.send_message(
                    message.chat.id,
                    'Это все локации выберите ту которую хотите изменить.',
                    reply_markup=buttons.Buttons(buttonsloc, 2)
                )
                bot.register_next_step_handler(message, editcont.locations1)
            else:
                bot.send_message(
                    message.chat.id,
                    'Для выбранного контрагента не указаны локации.',
                    reply_markup=buttons.Buttons(buttonsloc, 2)
                )
                bot.register_next_step_handler(message, editcont.locations1)
    # Сохранение имени новой локации
    def locations3(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        db.update_records(
            'Locations',
            ['name'],
            [message.text],
            'id',
            ActiveUser[message.chat.id]['curlocation']
        )
        bot.send_message(
            message.chat.id,
            'Название изменено.\nЧто вы хотите изменить?',
            reply_markup=buttons.Buttons(['Изменить локацию', 'Изменить название', '🗑️ Удалить', '🚫 Отмена'], 3)
        )
        bot.register_next_step_handler(message, editcont.locations2)
    # Удаление локации
    def locations4(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        if message.text == '✅ Да':
            db.delete_record('Locations', 'id', ActiveUser[message.chat.id]['curlocation'])
            bot.send_message(
                message.chat.id,
                'Локация удалена.',
                reply_markup=buttons.clearbuttons()
            )
            editcontragent(message)
            bot.register_next_step_handler(message, editcont.ec2)
        elif message.text == '⛔️ Нет':
            bot.send_message(
                message.chat.id,
                'Удаление отменено.\nЧто вы хотите изменить?',
                reply_markup=buttons.Buttons(['Изменить локацию', 'Изменить название', '🗑️ Удалить', '🚫 Отмена'], 3)
            )
            bot.register_next_step_handler(message, editcont.locations2)
        else:
            locationtodelete = db.get_record_by_id('Locations', ActiveUser[message.chat.id]['curlocation'])[2]
            Contrlocation = db.get_record_by_id('Contragents', ActiveUser[message.chat.id]['inn'])[1]
            bot.send_message(
                message.chat.id,
                f'Вы не подтвердили удаление.\nУдалить локацию {locationtodelete} у {Contrlocation}?',
                reply_markup=buttons.Buttons(['✅ Да','⛔️ Нет'])
            )
            bot.register_next_step_handler(message, editcont.locations4)
# основная форма рдактирования контрагента
def editcontragent(message):
    try:
        bot.delete_message(chat_id=ActiveUser[message.chat.id]['edcon'].chat.id, message_id=ActiveUser[message.chat.id]['edcon'].message_id)
    except:
        pass
    mess = "НАИМЕНОВАНИЕ:\n" + str(ActiveUser[message.chat.id]['contnew'][1])
    mess = mess + '\n\n' + "ИНН:\n" + str(ActiveUser[message.chat.id]['contnew'][0])
    if ActiveUser[message.chat.id]['contnew'][5] == 1:
        mess = mess + '\n\n' + "ТИП:\n" + 'Разовый'
    elif ActiveUser[message.chat.id]['contnew'][5] == 2:
        mess = mess + '\n\n' + "ТИП:\n" + 'Долгосрочный'
    elif ActiveUser[message.chat.id]['contnew'][5] == 3:
        mess = mess + '\n\n' + "ТИП:\n" + 'Физ лицо'
    else:
        mess = mess + '\n\n' + "ТИП:\n" + 'Не указан'
    mess = mess + '\n\n' + "АДРЕС:\n" + str(ActiveUser[message.chat.id]['contnew'][2])
    mess = mess + '\n\n' + "КОН'🆔 ИНН'ТАКТНОЕ ЛИЦО:\n" + str(ActiveUser[message.chat.id]['contnew'][3])
    mess = mess + '\n\n' + "ТЕЛЕФОН:\n" + str(ActiveUser[message.chat.id]['contnew'][4])
    mess = mess + '\n\n' + "ДОГОВОР:\n" + str(ActiveUser[message.chat.id]['contnew'][6])
    mess = mess + '\n\nЧТО ВЫ ХОТИТЕ ИЗМЕНИТЬ?'
    ActiveUser[message.chat.id]['edcon'] = bot.send_message(
        message.chat.id,
        mess,
        reply_markup=buttons.Buttons(['🆔 ИНН', '🏢 НАИМЕНОВАНИЕ', '🏷️ ТИП', '🛣️ АДРЕС', '📍 ЛОКАЦИИ', '🙋‍♂️ КОНТАКТНОЕ ЛИЦО', '📞 ТЕЛЕФОН', '📄 ДОГОВОР', '💾 Сохранить', '🚫 Отмена'], 3)
    )
    return
# Новая заявка
class NewTask:
    # Поиск контрагента по ИНН
    def nt1(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        ActiveUser[message.chat.id]['added'] = datetime.now().strftime("%d.%m.%Y %H:%M")
        ActiveUser[message.chat.id]['manager'] = message.chat.id
        ActiveUser[message.chat.id]['status'] = 1
        if message.text == '🚫 Отмена':
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.delete_message(chat_id=ActiveUser[message.chat.id]['sentmes'].chat.id, message_id=ActiveUser[message.chat.id]['sentmes'].message_id)
            bot.send_message(
                message.chat.id,
                'Выберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.register_next_step_handler(message, MainMenu.Main2)
        elif message.text.split()[0].isdigit():
            processing = bot.send_message(message.chat.id, '⏳')
            if message.text.split()[0].isdigit():
                inn = message.text.split()[0]
            else:
                inn = message.text.split()[1]
            ActiveUser[message.chat.id]['inn'] = inn
            findcont = db.get_record_by_id('Contragents', inn)
            if findcont == None:
                bot.delete_message(chat_id=message.chat.id, message_id=processing.message_id)
                bot.send_message(
                    message.chat.id,
                    'Контрагент с указанным Вами ИНН не найден. \nБудет добавлен новый.\nВыберите тип клиента',
                    reply_markup=buttons.Buttons(['Разовый', 'Долгосрочный', 'Физ. лицо'])
                )
                bot.register_next_step_handler(message, NewTask.NeContr1)
            else:
                bot.delete_message(chat_id=message.chat.id, message_id=processing.message_id)
                client = db.get_record_by_id('Contragents', inn)
                if client[5] is not None and ActiveUser[message.chat.id]['nt'] == 1:
                    bot.send_message(
                        message.chat.id,
                        'Выбран клиент - ' + str(client[1]) + '\nЗаявка или техника?',
                        reply_markup=buttons.Buttons(['📝 Заявка','🖨️ Техника'])
                    )
                    bot.register_next_step_handler(message, NewTask.tech1)
                elif ActiveUser[message.chat.id]['nt'] == 0:
                    ActiveUser[message.chat.id]['changecontrintask'] = inn
                    bot.send_message(
                        message.chat.id,
                        f'Контрагент заявки будет изменен на {str(client[1])}',
                        reply_markup=buttons.Buttons(['✅ Да', '⛔️ Нет'])
                    )
                    bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
                    bot.register_next_step_handler(message, Task.task6)
                else:
                    bot.send_message(
                        message.chat.id,
                        'У выбранного клиента - ' + str(client[1]) + ' не определен тип и договор.\nПожалуйста выберите тип клиента.',
                        reply_markup=buttons.Buttons(['Разовый', 'Долгосрочный', 'Физ. лицо'])
                    )
                    bot.register_next_step_handler(message, NewTask.type1)
        else:
            processing = bot.send_message(message.chat.id, '⏳')
            contrs = db.select_table('Contragents')
            res = functions.search_items(message.text, contrs)
            contbuttons = []
            contbuttons.append('🚫 Отмена')
            if len(res) > 0:
                for i in res:
                    line = str(i[0]) + ' ' + str(i[1])
                    if len(contbuttons) < 20:
                        contbuttons.append(line)
                bot.delete_message(chat_id=message.chat.id, message_id=processing.message_id)
                try:
                    bot.send_message(
                        message.chat.id,
                        'Если нужный клиент не вышел в списке, попробуйте перефразировать и ввести снова.\nВыберите контрагента из списка, или введите его ИНН, ПИНФЛ, серию пасспорта или повторите поиск.',
                        reply_markup=buttons.Buttons(contbuttons, 1)
                    )
                except Exception as e:
                    logging.error(e)
                    pass
            else:
                bot.delete_message(chat_id=message.chat.id, message_id=processing.message_id)
                bot.send_message(
                    message.chat.id,
                    'Контрагент не найден.',
                    reply_markup=buttons.Buttons(contbuttons, 1)
                )
            bot.register_next_step_handler(message, NewTask.nt1)
    # Техника
    def tech1(message):
        global ActiveUser
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        if message.text == '📝 Заявка':
            bot.send_message(
                message.chat.id,
                'опишите проблему клиента...',
                reply_markup=buttons.clearbuttons()
            )
            bot.register_next_step_handler(message, NewTask.ntlocation1)
        elif message.text == '🖨️ Техника':
            ActiveUser[message.chat.id]['status'] = 5
            bot.send_message(
                message.chat.id,
                'Укажите модель и производителя...',
                reply_markup=buttons.clearbuttons()
            )
            bot.register_next_step_handler(message, NewTask.tech2)
    def tech2(message):
        global ActiveUser
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        ActiveUser[message.chat.id]['task'] = message.text
        bot.send_message(
            message.chat.id,
            'Опишите проблему с техникой...',
            reply_markup=buttons.clearbuttons()
        )
        bot.register_next_step_handler(message, NewTask.tech3)
    def tech3(message):
        global ActiveUser
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        ActiveUser[message.chat.id]['task'] = ActiveUser[message.chat.id]['task'] + '\n======================\n' + message.text
        bot.send_message(
            message.chat.id,
            'Укажите номер квитанции и дату...\nПример:\n№ 10 от 01.01.2023',
            reply_markup=buttons.clearbuttons()
        )
        bot.register_next_step_handler(message, NewTask.nt2)
    # Тип соглашения контрагента если не добавлен в реквизиты контрагента
    def type1(message):
        global ActiveUser
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        if message.text == 'Разовый':
            ActiveUser[message.chat.id]['ds'] = 1
            bot.send_message(
                message.chat.id,
                'Кратко опишите проблему клиента',
                reply_markup=buttons.clearbuttons()
            )
            bot.register_next_step_handler(message, NewTask.ntlocation1)
        elif message.text == 'Долгосрочный':
            ActiveUser[message.chat.id]['ds'] = 2
            bot.send_message(
                message.chat.id,
                'Укажите номер и дату договора.',
                reply_markup=buttons.clearbuttons()
            )
            bot.register_next_step_handler(message, NewTask.type2)
        elif message.text == 'Физ. лицо':
            ActiveUser[message.chat.id]['ds'] = 3
            bot.send_message(
                message.chat.id,
                'Кратко опишите проблему клиента',
                reply_markup=buttons.clearbuttons()
            )
            bot.register_next_step_handler(message, NewTask.ntlocation1)
        else:
            bot.send_message(
                message.chat.id,
                'Ошибка ввода!\nВыберите тип клиента.',
                reply_markup=buttons.Buttons(['Разовый', 'Долгосрочный', 'Физ. лицо'])
            )
            bot.register_next_step_handler(message, NewTask.type1)
    def type2(message):
        global ActiveUser
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        ActiveUser[message.chat.id]['contract'] = message.text
        db.update_records(
            "Contragents",
            ["ds", "contract"],
            [ActiveUser[message.chat.id]['ds'], ActiveUser[message.chat.id]['contract']],
            "id", ActiveUser[message.chat.id]['inn']
        )
        bot.send_message(
            message.chat.id,
            'Заявка или техника?',
            reply_markup=buttons.Buttons(['📝 Заявка','🖨️ Техника'])
        )
        bot.register_next_step_handler(message, NewTask.tech1)
    # Обработка ошибки ввода ИНН
    def innerror(message):
        global ActiveUser
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        if message.text == 'Ввести снова':
            contragents = db.select_table('Contragents', ['id', 'cname'])
            bot.send_message(
                message.chat.id,
                'Выберите клиента или введите его ИНН.',
                reply_markup=buttons.Buttons(functions.listgen(contragents, [0, 1], 2))
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, NewTask.nt1)
        elif message.text == '🏠 Главное меню':
            ActiveUser[message.chat.id].clear()
            bot.send_message(
                message.chat.id,
                'Добро пожаловать в систему. Выберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, MainMenu.Main2)
    # Добавление нового контрагента
    def NeContr1(message):
        global ActiveUser
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        if message.text == 'Разовый':
            ActiveUser[message.chat.id]['ds'] = 1
            ActiveUser[message.chat.id]['contract'] = '...'
            bot.send_message(
                message.chat.id,
                'Введите наименование организации.',
                reply_markup=buttons.clearbuttons()
            )
            bot.register_next_step_handler(message, NewTask.NeContr3)
        elif message.text == 'Долгосрочный':
            ActiveUser[message.chat.id]['ds'] = 2
            bot.send_message(
                message.chat.id,
                'Укажите номер и дату договора..',
                reply_markup=buttons.clearbuttons()
            )
            bot.register_next_step_handler(message, NewTask.NeContr2)
        elif message.text == 'Физ. лицо':
            ActiveUser[message.chat.id]['ds'] = 3
            ActiveUser[message.chat.id]['contract'] = '...'
            bot.send_message(
                message.chat.id,
                'Введите Ф.И.О. клиента.',
                reply_markup=buttons.clearbuttons()
            )
            bot.register_next_step_handler(message, NewTask.NeContr3)
        else:
            bot.send_message(
                message.chat.id,
                'Ошибка ввода!\nВыберите тип клиента.',
                reply_markup=buttons.Buttons(['Разовый', 'Долгосрочный', 'Физ. лицо'])
            )
            bot.register_next_step_handler(message, NewTask.NeContr1)
    def NeContr2(message):
        global ActiveUser
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        ActiveUser[message.chat.id]['contract'] = message.text
        bot.send_message(
            message.chat.id,
            'Введите наименование организации.',
            reply_markup=buttons.clearbuttons()
        )
        bot.register_next_step_handler(message, NewTask.NeContr3)
    def NeContr3(message):
        global ActiveUser
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        ActiveUser[message.chat.id]['cname'] = message.text
        bot.send_message(
            message.chat.id,
            'Укажите адрес клиента или отправьте локацию.',
            reply_markup=buttons.clearbuttons()
        )

        if ActiveUser[message.chat.id]['ds'] == 3:
            bot.register_next_step_handler(message, NeContr5)

        else:
            bot.register_next_step_handler(message, NeContr4)
    def NeContr6(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        ActiveUser[message.chat.id]['cphone'] = message.text
        ActiveUser[message.chat.id]['mess'] = 'Подтвердите информацию:\n\n'
        if ActiveUser[message.chat.id]['ds'] == 1:
            ActiveUser[message.chat.id]['mess'] = ActiveUser[message.chat.id]['mess'] + 'Разовый\n'
            ActiveUser[message.chat.id]['mess'] = ActiveUser[message.chat.id]['mess'] + '\nНаименование организации: ' + ActiveUser[message.chat.id]['cname']
            ActiveUser[message.chat.id]['mess'] = ActiveUser[message.chat.id]['mess'] + '\nКонтактное лицо: ' + ActiveUser[message.chat.id]['cperson']
        elif ActiveUser[message.chat.id]['ds'] == 2:
            ActiveUser[message.chat.id]['mess'] = ActiveUser[message.chat.id]['mess'] + 'Долгосрочный\n'
            ActiveUser[message.chat.id]['mess'] = ActiveUser[message.chat.id]['mess'] + '\nНаименование организации: ' + ActiveUser[message.chat.id]['cname']
            ActiveUser[message.chat.id]['mess'] = ActiveUser[message.chat.id]['mess'] + '\nДоговор: ' + ActiveUser[message.chat.id]['contract']
            ActiveUser[message.chat.id]['mess'] = ActiveUser[message.chat.id]['mess'] + '\nКонтактное лицо: ' + ActiveUser[message.chat.id]['cperson']
        elif ActiveUser[message.chat.id]['ds'] == 3:
            ActiveUser[message.chat.id]['mess'] = ActiveUser[message.chat.id]['mess'] + 'Физ. лицо\n'
            ActiveUser[message.chat.id]['mess'] = ActiveUser[message.chat.id]['mess'] + '\nФ И О: ' + ActiveUser[message.chat.id]['cname']
        ActiveUser[message.chat.id]['mess'] = ActiveUser[message.chat.id]['mess'] + '\nАдрес: ' + ActiveUser[message.chat.id]['cadr']
        ActiveUser[message.chat.id]['mess'] = ActiveUser[message.chat.id]['mess'] + '\nТелефон: ' + ActiveUser[message.chat.id]['cphone']
        bot.send_message(
            message.chat.id,
            ActiveUser[message.chat.id]['mess'],
            reply_markup=buttons.Buttons(['✅ Да', '⛔️ Нет'])
        )
        bot.register_next_step_handler(message, NewTask.NeContr7)
    def NeContr7(message):
        if message.text == '✅ Да':
            username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
            logging.info(f'{username} Отправил запрос - {message.text}')
            global ActiveUser
            contragent = [
                ActiveUser[message.chat.id]['inn'],
                ActiveUser[message.chat.id]['cname'],
                ActiveUser[message.chat.id]['cadr'],
                ActiveUser[message.chat.id]['cperson'],
                ActiveUser[message.chat.id]['cphone'],
                ActiveUser[message.chat.id]['ds'],
                ActiveUser[message.chat.id]['contract']
            ]
            db.insert_record('Contragents', contragent)
            if ActiveUser[message.chat.id]['nt'] == 0:
                ActiveUser[message.chat.id]['changecontrintask'] = ActiveUser[message.chat.id]['inn']
                client = db.get_record_by_id('Contragents', ActiveUser[message.chat.id]['changecontrintask'])
                bot.send_message(
                    message.chat.id,
                    f'Контрагент заявки будет изменен на {str(client[1])}',
                    reply_markup=buttons.Buttons(['✅ Да', '⛔️ Нет'])
                )
                bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
                bot.register_next_step_handler(message, Task.task6)
            elif ActiveUser[message.chat.id]['nt'] == 1:
                bot.send_message(
                    message.chat.id,
                    'Заявка или техника?',
                    reply_markup=buttons.Buttons(['📝 Заявка','🖨️ Техника'])
                )
                bot.register_next_step_handler(message, NewTask.tech1)
        elif message.text == '⛔️ Нет':
            bot.send_message(
                message.chat.id,
                'Контрагент не добавлен.\nВыберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.register_next_step_handler(message, MainMenu.Main2)
        else:
            bot.send_message(
                message.chat.id,
                'Ошибка ввода!\n' + ActiveUser[message.chat.id]['mess'],
                reply_markup=buttons.Buttons(['✅ Да', '⛔️ Нет'])
            )
            bot.register_next_step_handler(message, NewTask.NeContr7)
    # Выбор локации
    def ntlocation1(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        if ActiveUser[message.chat.id]['status'] == 1:
            ActiveUser[message.chat.id]['task'] = message.text
        else:
            ActiveUser[message.chat.id]['task'] = ActiveUser[message.chat.id]['task'] + '\n======================\n' + message.text
        locations = db.select_table_with_filters('Locations', {'inn': ActiveUser[message.chat.id]['inn']})
        clocations = ['⏩ Пропустить']
        if len(locations) > 0:
            for i in locations:
                line = str(i[0]) + ' ' + str(i[2])
                clocations.append(line)
            clocations.append('🆕 Добавить филиал')
            bot.send_message(
                message.chat.id,
                'Выберите филиал, или введите название филиала',
                reply_markup=buttons.Buttons(clocations,2)
            )
        else:
            clocations.append('🆕 Добавить филиал')
            bot.send_message(
                message.chat.id,
                'У выбранного контрагента нет назначенных локаций!\nДобавьте новую.',
                reply_markup=buttons.Buttons(clocations)
            )
        bot.register_next_step_handler(message, NewTask.ntlocation2)
    # Проверка нажатия на кнопки выбора локаций
    def ntlocation2(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        if message.text == '⏩ Пропустить':
            ActiveUser[message.chat.id]['location'] = None
            conf(message)
            bot.register_next_step_handler(message, NewTask.nt3)
        elif message.text == '🆕 Добавить филиал':
            bot.send_message(
                message.chat.id,
                'Отправьте локацию',
                reply_markup=buttons.clearbuttons()
            )
            bot.register_next_step_handler(message, newlocationintask1)
        elif len(message.text.split()) > 0 and message.text.split()[0].isdigit():
            if message.text.split()[0].isdigit():
                ActiveUser[message.chat.id]['location'] = int(message.text.split()[0])
                conf(message)
                bot.register_next_step_handler(message, NewTask.nt3)
            else:
                bot.send_message(
                    message.chat.id,
                    'Ошибка вы не выбрали локацию.\nВыберите локацию или добавьте новую.',
                    reply_markup=buttons.clearbuttons()
                )
                bot.register_next_step_handler(message, NewTask.ntlocation1)
        else:
            contloc = db.select_table_with_filters('Locations', {'inn': ActiveUser[message.chat.id]['inn']})
            res = functions.search_items(message.text, contloc)
            but = []
            but.append('⏩ Пропустить')
            if len(res) > 0:
                for r in res:
                    line = str(r[0]) + ' ' + str(r[2])
                    but.append(line)
                but.append('🆕 Добавить филиал')
                bot.send_message(
                    message.chat.id,
                    'Выберите филиал, или введите название филиала',
                    reply_markup=buttons.Buttons(but, 2)
                )
                bot.register_next_step_handler(message, NewTask.ntlocation2)
            else:
                but.append('🆕 Добавить филиал')
                bot.send_message(
                    message.chat.id,
                    'Совпадений не найдено, попробуйте еще раз или добавьте филиал',
                    reply_markup=buttons.Buttons(but)
                )
                bot.register_next_step_handler(message, NewTask.ntlocation2)  
    # Добавление новой заявки в базу данных
    def nt2(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser, continue_polling
        ActiveUser[message.chat.id]['location'] = 999
        ActiveUser[message.chat.id]['task'] = ActiveUser[message.chat.id]['task'] + '\n======================\n' + message.text
        conf(message)
        bot.register_next_step_handler(message, NewTask.nt3)
    def nt3(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser, continue_polling
        processing = bot.send_message(message.chat.id, '⏳')
        if message.text == '✅ Да':
            task = [
                None,
                ActiveUser[message.chat.id]['added'],
                ActiveUser[message.chat.id]['manager'],
                ActiveUser[message.chat.id]['inn'],
                ActiveUser[message.chat.id]['task'],
                None,
                None,
                None,
                None,
                None,
                None,
                ActiveUser[message.chat.id]['status'],
                ActiveUser[message.chat.id]['location']
            ]
            db.insert_record('Tasks',task)
            tid = db.get_last_record('Tasks')[0]
            sendtoall(functions.curtask(tid), buttons.buttonsinline([['👍 Принять', 'confirm ' + str(tid)], ['📎 Назначить', 'set ' + str(tid)]]), message.chat.id, 1)
            bot.delete_message(chat_id=message.chat.id, message_id=processing.message_id)
            bot.send_message(
                message.chat.id,
                'Заявка успешно зарегистрирована.\nВыберрите операцию',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            continue_polling = True
            bot.register_next_step_handler(message, MainMenu.Main2)
        elif message.text == '⛔️ Нет':
            bot.send_message(
                message.chat.id,
                'Новая заявка удалена.\nВыберрите операцию',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.delete_message(chat_id=message.chat.id, message_id=processing.message_id)
            bot.register_next_step_handler(message, MainMenu.Main2)
        else:
            bot.send_message(
                message.chat.id,
                'Сначала подтвердите сохранение.\nСохранить заявку?',
                reply_markup=buttons.Buttons(['✅ Да', '⛔️ Нет'])
            )
            bot.delete_message(chat_id=message.chat.id, message_id=processing.message_id)
            bot.register_next_step_handler(message, NewTask.nt3)
# сообщение для подтверждения заявки
def conf(message):
    confmes = 'Подтвердите заявку. \nЗаявка от: '
    confmes = confmes + ActiveUser[message.chat.id]['added']
    record = db.get_record_by_id('Contragents', ActiveUser[message.chat.id]['inn'])
    if ActiveUser[message.chat.id]['location'] is not None:
        location = db.get_record_by_id('Locations', ActiveUser[message.chat.id]['location'])[2]
    else:
        location = ''
    confmes = confmes + '\nКлиент: ' + (record[1] if record[1] is not None else '') + (f" {location}" if ActiveUser[message.chat.id]['location'] is not None else '')
    confmes = confmes + '\nТекст заявки: ' + ActiveUser[message.chat.id]['task']
    confmes = confmes + '\nАдрес: ' + (record[2] if record[2] is not None else '')
    confmes = confmes + '\nКонтактное лицо: ' + (record[3] if record[3] is not None else '')
    confmes = confmes + '\nКонтактный номер: ' + (record[4] if record[4] is not None else '')
    bot.send_message(
        message.chat.id,
        confmes,
        reply_markup=buttons.Buttons(['✅ Да', '⛔️ Нет'])
    )
    return
# выбранная заявка и действия
class Task:

    def task1(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser, continue_polling
        if message.text == '👍 Принять':
            processing = bot.send_message(message.chat.id, '⏳')
            if db.get_record_by_id('Tasks', ActiveUser[message.chat.id]['task'])[11] == 5:
                stat = 6
            else:
                stat = 2
            if db.get_record_by_id('Tasks', ActiveUser[message.chat.id]['task'])[11] == 1 or db.get_record_by_id('Tasks', ActiveUser[message.chat.id]['task'])[11] == 5:
                db.update_records(
                    'Tasks',
                    [
                        'confirmed',
                        'master',
                        'status'
                    ], [
                        datetime.now().strftime("%d.%m.%Y %H:%M"),
                        message.chat.id, stat
                    ],
                    'id',
                    ActiveUser[message.chat.id]['task']
                )
                tk = functions.curtask(ActiveUser[message.chat.id]['task'])
                mes = str(db.get_record_by_id('Users', message.chat.id)[2]) + ' ' + str(db.get_record_by_id('Users', message.chat.id)[1]) + '\nПринял заявку:\n\n' + tk
                mark = ''
                exn = message.chat.id
                if sendedmessages is not None:
                    for line in sendedmessages:
                        try:
                            bot.delete_message(line[0], line[1])
                        except Exception as e:
                            logging.error(e)
                sendtoall(mes, mark, exn)
                bot.delete_message(chat_id=message.chat.id, message_id=processing.message_id)
                bot.send_message(
                    message.chat.id,
                    'Вы приняли заявку.\n\nВыберите операцию',
                    reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
                )
            else:
                bot.send_message(
                    message.chat.id,
                    "Вы не можете принять эту заявку!",
                    reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
                )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            continue_polling = True
        elif message.text == '🖊️ Дополнить':
            bot.send_message(
                message.chat.id,
                'Напишите что вы хотели дополнить...',
                reply_markup=buttons.clearbuttons()
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, Task.task5)
        elif message.text == '📎 Переназначить' or message.text == '📎 Назначить':
            users = db.select_table('Users')
            bot.send_message(
                message.chat.id,
                'Выберите мастера...',
                reply_markup=buttons.Buttons(functions.listgen(users, [0, 1, 2], 3), 1)
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, Task.task4)
        elif message.text == '✅ Выполнено':
            processing = bot.send_message(message.chat.id, '⏳')
            manager = str(db.get_record_by_id('Tasks', ActiveUser[message.chat.id]['task'])[6])
            if manager == str(message.chat.id):
                if db.get_record_by_id('Tasks', ActiveUser[message.chat.id]['task'])[11] == 2:
                    stat = 3
                else:
                    stat = 7
                db.update_records(
                    'Tasks',[
                        'done',
                        'status'
                    ],[
                        datetime.now().strftime("%d.%m.%Y %H:%M"),
                        stat
                    ],
                    'id',
                    ActiveUser[message.chat.id]['task']
                )
                tk = functions.curtask(ActiveUser[message.chat.id]['task'])
                mes = str(db.get_record_by_id('Users', message.chat.id)[2]) + ' ' + str(db.get_record_by_id('Users', message.chat.id)[1]) + '\nВыполнил заявку:\n\n' + tk
                mark = ''
                exn = message.chat.id
                sendtoall(mes, mark, exn)
                bot.delete_message(chat_id=message.chat.id, message_id=processing.message_id)
                bot.send_message(
                    message.chat.id,
                    'Вы завершили заявку.',
                    reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
                )
                continue_polling = True
        elif message.text == '🙅‍♂️ Отказаться от заявки':
            manager = str(db.get_record_by_id('Tasks', ActiveUser[message.chat.id]['task'])[6])
            if manager == str(message.chat.id):
                processing = bot.send_message(message.chat.id, '⏳')
                confdate = db.get_record_by_id('Tasks', ActiveUser[message.chat.id]['task'])[5]
                db.update_records(
                    'Tasks',
                    [
                        'more',
                        'master',
                        'status'
                    ], [
                        'Мастер ' + str(db.get_record_by_id('Users', message.chat.id)[2]) + ' ' + str(db.get_record_by_id('Users', message.chat.id)[1]) + ' принял заявку ' + str(confdate) + '.\n ' + str(datetime.now().strftime("%d.%m.%Y %H:%M")) + 'отказался от выполнения',
                        '',
                        1
                    ],
                    'id',
                    ActiveUser[message.chat.id]['task']
                )
                tk = functions.curtask(ActiveUser[message.chat.id]['task'])
                mes = 'Пользователь ' + str(db.get_record_by_id('Users', message.chat.id)[1]) + 'Отказался от заявки:\n\n' + tk
                mark = ''
                exn = message.chat.id
                sendtoall(mes, mark, exn)
                bot.delete_message(chat_id=message.chat.id, message_id=processing.message_id)
                bot.send_message(
                    message.chat.id,
                    'Вы отказались от заявки.',
                    reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
                )
                continue_polling = True
            else:
                master = db.get_record_by_id('Users', manager)[1]
                bot.send_message(
                    message.chat.id,
                    'Вы не можете отказаться от этой заявки, так как она не Ваша.\nЗаявку принял ' + str(master),
                    reply_markup=buttons.Buttons(['🏠 Главное меню'])
                )
                bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        elif message.text == '🚫 Отменить заявку':
            manager = str(db.get_record_by_id('Tasks', ActiveUser[message.chat.id]['task'])[2])
            bot.send_message(
                message.chat.id,
                'Вы уверены, что хотите отменить заявку?',
                reply_markup=buttons.Buttons(['✅ Да', '⛔️ Нет'])
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, Task.task2)
        elif message.text == '↩️ Назад':
            bot.send_message(
                message.chat.id,
                'Выберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            continue_polling = True
        elif message.text == '🤵 Изменить контрагента':
            bot.send_message(
                message.chat.id,
                'введите ИНН контрагента',
                reply_markup=buttons.clearbuttons()
            )
            ActiveUser[message.chat.id]['nt'] = 0
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, NewTask.nt1)
        elif message.text == '✏️ Изменить текст заявки':
            bot.send_message(
                message.chat.id,
                'Введите новый текст заявки.\n\n‼️ ВНИМАНИЕ ‼️\nУчтите что старый текст будет заменен новым поэтому скопируйте старый и отредактируйте.',
                reply_markup=buttons.clearbuttons()
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, Task.task7_1)
        elif message.text == '📍 Локация':
            print('📍 Локация')
            location = db.get_record_by_id('Locations', db.get_record_by_id('Tasks', ActiveUser[message.chat.id]['task'])[12])
            if location is not None:
                loc = types.Location(location[4], location[3])
                bot.send_location(message.chat.id, loc.latitude, loc.longitude)
                bot.send_message(
                    message.chat.id,
                    'Выберите операцию',
                    reply_markup=buttons.Buttons(['📍 Указать локацию', '🚫 Отмена'])
                )
                bot.register_next_step_handler(message, Task.locations1)
            else:
                bot.send_message(
                    message.chat.id,
                    'Прошу прощения но указанная локаия либо не была добавлена, или была удалена.\nВыберите операцию',
                    reply_markup=buttons.Buttons(['📍 Указать локацию', '🚫 Отмена'])
                )
                bot.register_next_step_handler(message, Task.locations1)
        else:
            bot.send_message(
                message.chat.id,
                'Ошибка ввода.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )

    def locations1(message):
        global ActiveUser, continue_polling
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        if message.text == '🚫 Отмена':
            print('Нажата отмена')
            bot.send_message(
                message.chat.id,
                'Выберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
        elif message.text == '📍 Указать локацию':
            inn = db.get_record_by_id('Tasks', ActiveUser[message.chat.id]['task'])[3]
            locations = db.select_table_with_filters('Locations', {'inn': inn})
            buttonsloc = []
            buttonsloc.append('🆕 Добавить филиал')
            if len(locations) > 0:
                for location in locations:
                    line = str(location[0]) + ' ' + str(location[2])
                    print(line)
                    buttonsloc.append(line)
            buttonsloc.append('🚫 Отмена')
            print(buttonsloc)
            bot.send_message(
                message.chat.id,
                'Выберите филиал',
                reply_markup=buttons.Buttons(buttonsloc, 2)
            )
            bot.register_next_step_handler(message, Task.locations2)

    def locations2(message):
        global ActiveUser, continue_polling
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        ActiveUser[message.chat.id]['inn'] = db.get_record_by_id('Tasks', ActiveUser[message.chat.id]['task'])[3]
        if message.text == '🆕 Добавить филиал':
            bot.send_message(
                message.chat.id,
                'Отправьте локацию.',
                reply_markup=buttons.clearbuttons()
            )
            bot.register_next_step_handler(message, tnl1)
        elif message.text == '🚫 Отмена':
            bot.send_message(
                message.chat.id,
                'Выберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            continue_polling = True
        elif message.text.split()[0].isdigit():
            selected = db.get_record_by_id('Locations', message.text.split()[0])
            db.update_records(
                'Tasks',
                ['location'],
                [selected[0]],
                'id',
                ActiveUser[message.chat.id]['task']
            )
            bot.send_message(
                message.chat.id,
                f'Выбрана локация {selected[2]}',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            continue_polling = True
        else:
            inn = db.get_record_by_id('Tasks', ActiveUser[message.chat.id]['task'])[3]
            locations = db.select_table_with_filters('Locations', {'inn': inn})
            buttonsloc = []
            buttonsloc.append('🆕 Добавить филиал')
            if len(locations) > 0:
                for location in locations:
                    buttonsloc.append(str(location[0]) + ' ' + str(location[2]))
            buttonsloc.append('🚫 Отмена')
            if len(locations) > 0:
                bot.send_message(
                    message.chat.id,
                    'Ошибка ввода!\nВыберите филиал',
                    reply_markup=buttons.Buttons(buttonsloc, 2)
                )
            bot.register_next_step_handler(message, Task.locations2)

    def task2(message):
        global ActiveUser, continue_polling
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        if message.text == '✅ Да':
            bot.send_message(
                message.chat.id,
                'Пожалуйста укажите причину отмены заявки.',
                reply_markup=buttons.clearbuttons()
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, Task.task3)
        elif message.text == '⛔️ Нет':
            bot.send_message(
                message.chat.id,
                'Выберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)

    def task3(message):
        global ActiveUser, continue_polling
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        processing = bot.send_message(message.chat.id, '⏳')
        db.update_records(
            'Tasks',[
                'canceled',
                'userc',
                'more',
                'status'
            ],[
                datetime.now().strftime("%d.%m.%Y %H:%M"),
                message.chat.id,
                message.text,
                4
            ],
            'id',
            ActiveUser[message.chat.id]['task']
        )
        tk = functions.curtask(ActiveUser[message.chat.id]['task'])
        mes = str(db.get_record_by_id('Users', message.chat.id)[2]) + ' ' + str(db.get_record_by_id('Users', message.chat.id)[1]) + '\nОтменил заявку:\n\n' + tk + '\n\nПРИЧИНА:\n' + message.text
        mark = ''
        exn = message.chat.id
        sendtoall(mes, mark, exn)
        bot.delete_message(chat_id=message.chat.id, message_id=processing.message_id)
        bot.send_message(
            message.chat.id,
            'Заявка отменена\n\nВыберите операцию.',
            reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
        )
        continue_polling = True

    def task4(message):
        global ActiveUser, continue_polling
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        if db.get_record_by_id('Tasks', ActiveUser[message.chat.id]['task'])[11] == 5:
            stat = 6
        else:
            stat = 2
        if message.text.split()[1] is None:
            users = db.select_table('Users')
            bot.send_message(
                message.chat.id,
                'Выберите мастера...',
                reply_markup=buttons.Buttons(functions.listgen(users, [0, 1, 2], 3), 1)
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, Task.task4)
        else:
            processing = bot.send_message(message.chat.id, '⏳')
            userm = message.text.split()[1]
            
            db.update_records(
                'Tasks',
                [
                    'confirmed',
                    'master',
                    'status'
                ], [
                    datetime.now().strftime("%d.%m.%Y %H:%M"),
                    userm, stat
                ],
                'id',
                ActiveUser[message.chat.id]['task']
            )
            if sendedmessages is not None:
                for line in sendedmessages:
                    try:
                        bot.delete_message(line[0], line[1])
                    except Exception as e:
                        logging.error(e)
            tk = functions.curtask(ActiveUser[message.chat.id]['task'])
            mes = str(db.get_record_by_id('Users', userm)[2]) + ' ' + str(db.get_record_by_id('Users', userm)[1]) + '\nбыл назначен исполнителем заявки:\n\n' + tk
            exn = message.chat.id
            sendtoall(mes, '', exn)
            bot.delete_message(chat_id=message.chat.id, message_id=processing.message_id)
            bot.send_message(
                message.chat.id,
                'Мастер назначен.\n\nВыберите операцию',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            continue_polling = True

    def task5(message):
        global ActiveUser, continue_polling
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        processing = bot.send_message(message.chat.id, '⏳')
        tasktext = db.get_record_by_id('Tasks', ActiveUser[message.chat.id]['task'])[4]
        db.update_records(
            'Tasks',
            ['task'],
            [tasktext + '\n\n ' + username + ' дополнил(а) заявку...\n' + message.text],
            'id',
            ActiveUser[message.chat.id]['task']
        )
        tk = functions.curtask(ActiveUser[message.chat.id]['task'])
        mes = str(db.get_record_by_id('Users', message.chat.id)[2]) + ' ' + str(db.get_record_by_id('Users', message.chat.id)[1]) + '\n дополнил(а) заявку:\n\n' + tk
        mark = ''
        exn = message.chat.id
        sendtoall(mes, mark, exn)
        if sendedmessages is not None:
            for line in sendedmessages:
                try:
                    bot.delete_message(line[0], line[1])
                except Exception as e:
                    logging.error(e)
        bot.delete_message(chat_id=message.chat.id, message_id=processing.message_id)
        bot.send_message(
            message.chat.id,
            'Заявка дополнена.\n\nВыберите операцию',
            reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
        )
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        continue_polling = True

    def task6(message):
        global ActiveUser, continue_polling
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        if message.text == '✅ Да':
            processing = bot.send_message(message.chat.id, '⏳')
            db.update_records(
                'Tasks',
                ['contragent'],
                [ActiveUser[message.chat.id]['changecontrintask']],
                'id',
                ActiveUser[message.chat.id]['task']
            )
            client = db.get_record_by_id('Contragents', ActiveUser[message.chat.id]['changecontrintask'])[1]
            tk = functions.curtask(ActiveUser[message.chat.id]['task'])
            mes = str(db.get_record_by_id('Users', message.chat.id)[2]) + ' ' + str(db.get_record_by_id('Users', message.chat.id)[1]) + '\n Изменил(а) клиента в заявке:\n\n' + tk
            mark = ''
            sendtoall(mes, mark, message.chat.id)
            bot.delete_message(chat_id=message.chat.id, message_id=processing.message_id)
            bot.send_message(
                message.chat.id,
                f'Клиент в заявке изменен на {client}.\n\nВыберите операцию',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            continue_polling = True
        elif message.text == '⛔️ Нет':
            bot.send_message(
                message.chat.id,
                'Выберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        else:
            bot.send_message(
                message.chat.id,
                'Неверная команда',
                reply_markup=buttons.Buttons(['✅ Да', '⛔️ Нет'])
            )
            bot.register_next_step_handler(message, Task.task6)

    def task7_1(message):
        global ActiveUser, continue_polling
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        taskt = db.get_record_by_id('Tasks', ActiveUser[message.chat.id]['task'])[4]
        ActiveUser[message.chat.id]['newtasktext'] = message.text
        bot.send_message(
            message.chat.id,
            f'Текст заявку будет изменен с:\n{taskt}\nНа:\n{message.text}\n\n Подтвердите информацию...',
            reply_markup=buttons.Buttons(['✅ Да','⛔️ Нет'])
        )
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        bot.register_next_step_handler(message, Task.task7_2)

    def task7_2(message):
        global ActiveUser, continue_polling
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        if message.text == '✅ Да':
            processing = bot.send_message(message.chat.id, '⏳')
            print(ActiveUser[message.chat.id]['newtasktext'])
            db.update_records(
                'Tasks',
                ['task'],
                [ActiveUser[message.chat.id]['newtasktext']],
                'id',
                ActiveUser[message.chat.id]['task']
            )
            tk = functions.curtask(ActiveUser[message.chat.id]['task'])
            mes = str(db.get_record_by_id('Users', message.chat.id)[2]) + ' ' + str(db.get_record_by_id('Users', message.chat.id)[1]) + '\n внес изменения в заявку\n\n' + tk
            mark = ''
            sendtoall(mes, mark, message.chat.id)
            bot.delete_message(chat_id=message.chat.id, message_id=processing.message_id)
            bot.send_message(
                message.chat.id,
                'Заявка успешно измненена.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            continue_polling = True
        elif message.text == '⛔️ Нет':
            bot.send_message(
                message.chat.id,
                'Выберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        else:
            bot.send_message(
                message.chat.id,
                'Вы не подтвердили информацию.\nЗаменить старый текст заявки на новый',
                reply_markup=buttons.Buttons(['✅ Да','⛔️ Нет'])
            )
            bot.register_next_step_handler(message, Task.task7_2)
# фильтр для запроса в базу
def filters(message):
    username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
    logging.info(f'{username} Отправил запрос - {message.text}')
    global ActiveUser
    messagetouser = 'По каким параметрам формировать список\n'
    if ActiveUser[message.chat.id]['filter']['from'] == '01.01.2000 00:00':
        messagetouser = messagetouser + '📅 Будут показаны все заявки за весь период.\n'
    else:
        messagetouser = messagetouser + '📅 Выбран период:\n c' + str(ActiveUser[message.chat.id]['filter']['from']) + ' по ' + str(ActiveUser[message.chat.id]['filter']['to']) + '\n'
    messagetouser = messagetouser + '\n📍 СТАТУС:\n'
    if ActiveUser[message.chat.id]['filter']['added'] == 1:
        messagetouser = messagetouser + '🔵 Зарегистрированные\n'
    if ActiveUser[message.chat.id]['filter']['confirmed'] == 1:
        messagetouser = messagetouser + '🟡 В работе\n'
    if ActiveUser[message.chat.id]['filter']['done'] == 1:
        messagetouser = messagetouser + '🟢 Завершенные\n'
    if ActiveUser[message.chat.id]['filter']['canceled'] == 1:
        messagetouser = messagetouser + '🔴 Отмененные'
    if ActiveUser[message.chat.id]['filter']['justmy'] == 1:
        messagetouser = messagetouser + '\n👤 Показать только мои заявки.'
    return messagetouser
# общий чат (пересылка сообщения всем пользователям)
class allchats:
    # пересылка пользовательского сообщения всем
    def chat1(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос в отправке всем пользователям - {message.text}')
        if message.text == '🏠 Главное меню' or message.text == '/start':
            logging.info('main')
            bot.send_message(
                message.chat.id,
                'Выберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, MainMenu.Main2)
        else:
            logging.info('message to all')
            users = db.select_table('Users')
            for user in users:
                try:
                    logging.info(f'sended message to user {user[2]} {user[1]}')
                    if user[0] != message.chat.id:
                        bot.forward_message(user[0], message.chat.id, message.message_id)
                except Exception as e:
                    logging.error(e)
                    pass
            # for user in users:
            #     try:
            #         pinm = bot.forward_message(user[0], message.chat.id, message.message_id)
            #         if message.from_user.id == 65241621 or message.from_user.id == 1669785252:
            #             bot.unpin_chat_message(user[0])
            #             bot.pin_chat_message(user[0], pinm.message_id)
            #         logging.info(f'sent message to user {user[3]} from {user[2]}')
            #     except Exception as e:
            #         logging.error(e)
            #         pass
            bot.register_next_step_handler(message, allchats.chat1)
# отчеты
class report:
    # Запрос в базу с параметрами
    def rep(message, daterep, dr = 1, conf = 0, added = 0, done = 0, canc = 0, master = 0, my = 0, tadded = 0, tconf = 0, tdone = 0):
        donetasks = []
        confirmedtasks = []
        addedtasks = []
        canceledtasks = []
        if tdone == 1:
            if my == 1:
                filt = {'status': 7, 'master': master}
            else:
                filt = {'status': 7}
            logging.info(f'Запрос в базу на технику у мастеров за {daterep}')
            tdonetasks = functions.listgen(db.select_table_with_filters('Tasks', filt, ['done'], [daterep+' 00:00'], [daterep+' 23:59']), [0, 1, 3, 4, 6], 1)
            if len(tdonetasks) != 0:
                bot.send_message(
                    message.chat.id,
                    '🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻\nГотовая техника.',
                    reply_markup=''
                )
                for line in tdonetasks:
                    taskid = line.split()[2]
                    bot.send_message(
                        message.chat.id,
                        line,
                        reply_markup=buttons.buttonsinline([['Показать подробности', 'tasklist '+taskid]])
                    )
        if tconf == 1:
            if my == 1:
                filt = {'status': 6, 'master': master}
            else:
                filt = {'status': 6}
            logging.info(f'Запрос в базу на технику у мастеров за {daterep}')
            tconftasks = functions.listgen(db.select_table_with_filters('Tasks', filt), [0, 1, 3, 4, 6], 1)
            if len(tconftasks) != 0:
                bot.send_message(
                    message.chat.id,
                    '🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻\nТехника у мастеров.',
                    reply_markup=''
                )
                for line in tconftasks:
                    taskid = line.split()[2]
                    bot.send_message(
                        message.chat.id,
                        line,
                        reply_markup=buttons.buttonsinline([['Показать подробности', 'tasklist '+taskid]])
                    )
        if tadded == 1:
            logging.info(f'Запрос в базу на принятую технику.')
            taddedtasks = functions.listgen(db.select_table_with_filters('Tasks', {'status': 5}), [0, 1, 3, 4, 6], 1)
            if len(taddedtasks) != 0:
                bot.send_message(
                    message.chat.id,
                    '🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻\nОжидает ремонта',
                    reply_markup=''
                )
                for line in taddedtasks:
                    taskid = line.split()[2]
                    bot.send_message(
                        message.chat.id,
                        line,
                        reply_markup=buttons.buttonsinline([['Показать подробности', 'tasklist '+taskid]])
                    )
        if done == 1:
            if my == 1:
                filt = {'status': 3, 'master': master}
            else:
                filt = {'status': 3}
            logging.info(f'Запрос в базу на выполненные заявки за {daterep}')
            donetasks = functions.listgen(db.select_table_with_filters('Tasks', filt, ['done'], [daterep+' 00:00'], [daterep+' 23:59']), [0, 1, 3, 4, 6], 1)
            if len(donetasks) != 0:
                bot.send_message(
                    message.chat.id,
                    '🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻\nВыполненные заявки',
                    reply_markup=''
                )
                for line in donetasks:
                    taskid = line.split()[2]
                    bot.send_message(
                        message.chat.id,
                        line,
                        reply_markup=buttons.buttonsinline([['Показать подробности', 'tasklist '+taskid]])
                    )
        if conf == 1:
            logging.info(f'Запрос в базу на принятые заявки за {daterep}')
            if master == 0:
                filt = {'status': 2}
            else:
                filt = {'status': 2, 'master': master}
            confirmedtasks = functions.listgen(db.select_table_with_filters('Tasks', filt), [0, 1, 3, 4, 6], 1)
            if master != 0 and len(confirmedtasks) == 0:
                bot.send_message(
                    message.chat.id,
                    'У вас нет заявок в работе.',
                    reply_markup=''
                )
            if len(confirmedtasks) != 0:
                bot.send_message(
                    message.chat.id,
                    '🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻\nЗаявки у мастеров',
                    reply_markup=''
                )
                for line in confirmedtasks:
                    taskid = line.split()[2]
                    bot.send_message(
                        message.chat.id,
                        line,
                        reply_markup=buttons.buttonsinline([['Показать подробности', 'tasklist '+taskid]])
                    )
        if added == 1:
            logging.info(f'Запрос в базу на зарегистрированные заявки за {daterep}')
            addedtasks = functions.listgen(db.select_table_with_filters('Tasks', {'status': 1}), [0, 1, 3, 4, 6], 1)
            if len(addedtasks) != 0:
                bot.send_message(
                    message.chat.id,
                    '🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻\nНе принятые заявки',
                    reply_markup=''
                )
                for line in addedtasks:
                    taskid = line.split()[2]
                    bot.send_message(
                        message.chat.id,
                        line,
                        reply_markup=buttons.buttonsinline([['Показать подробности', 'tasklist '+taskid]])
                    )
        if canc == 1:
            logging.info(f'Запрос в базу на отмененные заявки за {daterep}')
            canceledtasks = functions.listgen(db.select_table_with_filters('Tasks', {'status': 4}, ['canceled'], [daterep+' 00:00'], [daterep+' 23:59']), [0, 1, 3, 4, 6], 1)
            if len(canceledtasks) != 0:
                bot.send_message(
                    message.chat.id,
                    '🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻\nОтмененные',
                    reply_markup=''
                )
                for line in canceledtasks:
                    taskid = line.split()[2]
                    bot.send_message(
                        message.chat.id,
                        line,
                        reply_markup=buttons.buttonsinline([['Показать подробности', 'tasklist '+taskid]])
                    )
        if dr == 1:
            reports = '\nВыполнено - ' + str(len(donetasks)) + '\nНе распределенных - ' + str(len(addedtasks)) + '\nВ работе у мастеров - ' + str(len(confirmedtasks)) + '\nОтменено - ' + str(len(canceledtasks))
            reports = reports + '\n\nКоличество заявок выполненных мастерами:\n\n'
            users = db.select_table('Users')
            usersrep = []
            print(daterep)
            for i in users:
                tasks = len(db.select_table_with_filters('Tasks', {'master': i[0], 'status': 3}, ['done'], [str(daterep)+' 00:00'], [str(daterep)+' 23:59']))
                usersrep.append([i[2] + ' ' + i[1], tasks])
            sorted_usersrep = sorted(usersrep, key=lambda x: x[1], reverse=True)
            for j in sorted_usersrep:
                if j[1] != 0:
                    reports = reports + '\n' + j[0] + ' - ' + str(j[1])
            bot.send_message(
                message.chat.id,
                'ИТОГИ ДНЯ\n🔺🔺🔺🔺🔺🔺🔺🔺🔺🔺🔺🔺' + reports,
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
        else:
            if len(addedtasks) != 0 and len(confirmedtasks) != 0 and len(donetasks) != 0 and len(canceledtasks) != 0:
                bot.send_message(
                    message.chat.id,
                    'Выберите операцию.',
                    reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
                )
        bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
        bot.register_next_step_handler(message, MainMenu.Main2)
    # Реакия на нажатие кнопок меню отчетов
    def reportall(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        if message.text == '📋 Заявки у мастеров':
            logging.info('план отправлен.')
            users = db.select_table('Users')
            res = ''
            bot.send_message(
                message.chat.id,
                'ЗАЯВКИ У МАСТЕРОВ\n\n🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻',
                reply_markup=buttons.clearbuttons()
            )
            processing = bot.send_message(message.chat.id, '⏳')
            tc = 0
            for u in users:
                userid = u[0]
                daterep = str(datetime.now().strftime("%d.%m.%Y"))
                confirmed = db.select_table_with_filters('Tasks', {'status': 2, 'master': userid})
                done = db.select_table_with_filters('Tasks', {'status': 3, 'master': userid}, ['done'], [daterep+' 00:00'], [daterep+' 23:59'])
                canceled = db.select_table_with_filters('Tasks', {'status': 4, 'master': userid}, ['canceled'], [daterep+' 00:00'], [daterep+' 23:59'])
                if len(confirmed) > 0 or len(done) > 0 or len(canceled) > 0:
                    res = res + f'\n🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻\n👤 {u[2]} {u[1]}\n\nЗАЯВКИ В РАБОТЕ:\n'
                    if len(confirmed) > 0:
                        for i in confirmed:
                            contr = db.get_record_by_id('Contragents', i[3])[1]
                            adr = db.get_record_by_id('Contragents', i[3])[2]
                            res = res + f'\n🟡 - №{i[0]} от {i[1]} | {contr}\n{adr}'
                    if len(done) > 0:
                        for j in done:
                            contr = db.get_record_by_id('Contragents', j[3])[1]
                            adr = db.get_record_by_id('Contragents', j[3])[2]
                            res = res + f'\n🟢 - №{j[0]} от {j[1]} | {contr}\n{adr}'
                    if len(canceled) > 0:
                        for k in canceled:
                            contr = db.get_record_by_id('Contragents', k[3])[1]
                            adr = db.get_record_by_id('Contragents', k[3])[2]
                            res = res + f'\n🔴 - №{k[0]} от {k[1]} | {contr}\n{adr}'
                    bot.send_message(
                        message.chat.id,
                        res,
                        reply_markup=buttons.clearbuttons()
                    )
                    res = ''
                    tc = tc + 1
            bot.delete_message(chat_id=message.chat.id, message_id=processing.message_id)
            if tc == 0:
                bot.send_message(
                    message.chat.id,
                    'У мастеров нет заявок.\nВыберите действие',
                    reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
                )
            else:
                bot.send_message(
                    message.chat.id,
                    'Выберите действие',
                    reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
                )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, MainMenu.Main2)
        elif message.text == '🖨️ Техника у мастеров':
            logging.info('план отправлен.')
            users = db.select_table('Users')
            res = ''
            bot.send_message(
                message.chat.id,
                'ТЕХНИКА В РАБОТЕ\n\n🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻',
                reply_markup=buttons.clearbuttons()
            )
            processing = bot.send_message(message.chat.id, '⏳')
            tc = 0
            for u in users:
                userid = u[0]
                daterep = str(datetime.now().strftime("%d.%m.%Y"))
                confirmed = db.select_table_with_filters('Tasks', {'status': 6, 'master': userid})
                done = db.select_table_with_filters('Tasks', {'status': 7, 'master': userid}, ['done'], [daterep+' 00:00'], [daterep+' 23:59'])
                if len(confirmed) > 0 or len(done) > 0:
                    res = res + f'\n🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻\n👤 {u[2]} {u[1]}\n\nТЕХНИКА НА РЕМОНТЕ:\n'
                    if len(confirmed) > 0:
                        for i in confirmed:
                            tech = i[4].split('\n======================\n')[0] + '\n' + i[4].split('\n======================\n')[1]
                            res = res + f'\n🟨 - №{i[0]} от {i[1]}\n{tech}'
                    if len(done) > 0:
                        for j in done:
                            tech = j[4].split('\n======================\n')[0] + '\n' + j[4].split('\n======================\n')[1]
                            res = res + f'\n🟩 - №{j[0]} от {j[1]}\n{tech}'
                    bot.send_message(
                        message.chat.id,
                        res,
                        reply_markup=buttons.clearbuttons()
                    )
                    res = ''
                tc = tc + 1
            bot.delete_message(chat_id=message.chat.id, message_id=processing.message_id)
            if tc == 0:
                bot.send_message(
                    message.chat.id,
                    'У мастеров нет техники на ремонте.\nВыберите действие',
                    reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
                )
            else:
                bot.send_message(
                    message.chat.id,
                    'Выберите действие',
                    reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
                )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, MainMenu.Main2)
        elif message.text == '📊 Итоги дня':
            bot.send_message(
                message.chat.id,
                'Какой день вы хотите увидеть?',
                reply_markup = buttons.Buttons(['🌞 Сегодня', '🗓️ Другой день'])
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, report.reportall1)
        elif message.text == '📆 За период':
            bot.send_message(
                message.chat.id,
                'Укажите начало периода в формате:\nПРИМЕР: 01.01.2023 или 01,01,2023',
                reply_markup = buttons.clearbuttons()
            )
            bot.register_next_step_handler(message, report.period1)
        elif message.text == '🚫 Отмена':
            bot.send_message(
                message.chat.id,
                'Выберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.register_next_step_handler(message, MainMenu.Main2)
        else:
            bot.send_message(
                message.chat.id,
                'Не верная команда!\nВыберите какой отчет Вам нужен.',
                reply_markup=buttons.Buttons(['📋 Заявки у мастеров', '🖨️ Техника у мастеров', '📊 Итоги дня', '🚫 Отмена'])
            )
            bot.register_next_step_handler(message, report.reportall)
    # period
    def period1(message):# с
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        m1 = message.text
        m1 = m1.replace(' ', '.')
        m1 = m1.replace(',', '.')
        m = m1.split('.')
        if len(m[0]) == 2 and len(m[1]) == 2 and len(m[2]) == 4 and len(m) == 3:
            ActiveUser[message.chat.id]['daterepf'] = m1
            bot.send_message(
                message.chat.id,
                'Укажите конец периода в формате:\nПРИМЕР: 01.01.2023 или 01,01,2023',
                reply_markup=buttons.clearbuttons()
            )
            bot.register_next_step_handler(message, report.period2)
        else:
            bot.send_message(
                message.chat.id,
                'Не верный формат даты...\nУкажите начало периода в формате:\nПРИМЕР: 01.01.2023 или 01,01,2023',
                reply_markup=buttons.clearbuttons()
            )
            bot.register_next_step_handler(message, report.period1)
    def period2(message):# по
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        m1 = message.text
        m1 = m1.replace(' ', '.')
        m1 = m1.replace(',', '.')
        m = m1.split('.')
        if len(m[0]) == 2 and len(m[1]) == 2 and len(m[2]) == 4 and len(m) == 3:
            ActiveUser[message.chat.id]['daterept'] = m1
            fr = ActiveUser[message.chat.id]['daterepf']
            t = ActiveUser[message.chat.id]['daterept']
            bot.send_message(
                message.chat.id,
                f'Выбран период с {fr} по {t}\nКакой отчет вывести?',
                reply_markup=buttons.Buttons(['Все заявки','по мастерам','по клиенту'])
            )
            bot.register_next_step_handler(message, report.period3)
        else:
            bot.send_message(
                message.chat.id,
                'Не верный формат даты...\nУкажите конец периода в формате:\nПРИМЕР: 01.01.2023 или 01,01,2023',
                reply_markup=buttons.clearbuttons()
            )
            bot.register_next_step_handler(message, report.period2)
    def period3(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        if message.text == 'Все заявки':
            fr = ActiveUser[message.chat.id]['daterepf']
            t = ActiveUser[message.chat.id]['daterept']
            rept = db.select_table_with_filters('Tasks', {'status': 3}, ['done'], [fr+' 00:00'], [t+' 23:59'])
            sendrepfile(message, rept)
            bot.send_message(
                message.chat.id,
                'Выберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.register_next_step_handler(message, MainMenu.Main2)
        elif message.text == 'по мастерам':
            users = db.select_table('Users')
            btn = []
            for user in users:
                line = str(user[0]) + ' ' + str(user[2]) + ' ' + str(user[1])
                btn.append(line)
            bot.send_message(
                message.chat.id,
                'Выберите мастера.',
                reply_markup=buttons.Buttons(btn,1)
            )
            bot.register_next_step_handler(message, report.period4)
        elif message.text == 'по клиенту':
            bot.send_message(
                message.chat.id,
                'Введите ИНН, ПИНФЛ или серию пвсспоррта клиента.\nТак же Вы можете попытаться поискать контрагента по наименованию или его части\nНапримар:\nmonohrom\nВыдаст все компании из базы бота у которых в названии есть monohrom',
                reply_markup=buttons.Buttons(['🚫 Отмена'])
            )
            bot.register_next_step_handler(message, report.period5)

    def period4(message): # по мастерам
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        uid = message.text.split()[0]
        selecteduser = db.get_record_by_id('Users', uid)
        if uid.isdigit() and selecteduser != None:
            fr = ActiveUser[message.chat.id]['daterepf']
            t = ActiveUser[message.chat.id]['daterept']
            rept = db.select_table_with_filters('Tasks', {'master': selecteduser[0], 'status': 3}, ['done'], [fr+' 00:00'], [t+' 23:59'])
            sendrepfile(message, rept)
            bot.send_message(
                message.chat.id,
                'Выберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.register_next_step_handler(message, MainMenu.Main2)
        else:
            users = db.select_table('Users')
            btn = []
            for user in users:
                line = str(user[0]) + ' ' + str(user[2]) + ' ' + str(user[1])
                btn.append(line)
            bot.send_message(
                message.chat.id,
                'Выберите мастера.',
                reply_markup=buttons.Buttons(btn,1)
            )
            bot.register_next_step_handler(message, report.period4)
    def period5(message): # по клиентам
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        global ActiveUser
        if message.text == '🚫 Отмена':
            bot.send_message(
                message.chat.id,
                'Выберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.register_next_step_handler(message, MainMenu.Main2)
        elif message.text.split()[0].isdigit():
            inn = int(message.text.split()[0])
            client = db.get_record_by_id('Contragents', inn)
            if client[5] is not None:
                fr = ActiveUser[message.chat.id]['daterepf']
                t = ActiveUser[message.chat.id]['daterept']
                rept = db.select_table_with_filters('Tasks', {'contragent': inn, 'status': 3}, ['done'], [fr+' 00:00'], [t+' 23:59'])
                sendrepfile(message, rept)
                bot.send_message(
                    message.chat.id,
                    'Выберите операцию.',
                    reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
                )
                bot.register_next_step_handler(message, MainMenu.Main2)
            else:
                bot.send_message(
                    message.chat.id,
                    'Контрагент не найден.',
                    reply_markup=buttons.Buttons(['🚫 Отмена'])
                )
                bot.register_next_step_handler(message, report.period5)
        else:
            contrs = db.select_table('Contragents')
            res = functions.search_items(message.text, contrs)
            contbuttons = []
            contbuttons.append('🚫 Отмена')
            if len(res) > 0:
                for i in res:
                    line = str(i[0]) + ' ' + str(i[1])
                    if len(contbuttons) < 20:
                        contbuttons.append(line)
                try:
                    bot.send_message(
                        message.chat.id,
                        'Если нужный клиент не вышел в списке, попробуйте перефразировать и ввести снова.\nВыберите контрагента из списка, или введите его ИНН, ПИНФЛ, серию пасспорта или повторите поиск.',
                        reply_markup=buttons.Buttons(contbuttons, 1)
                    )
                except Exception as e:
                    logging.error(e)
                    pass
            else:
                bot.send_message(
                    message.chat.id,
                    'Контрагент не найден.',
                    reply_markup=buttons.Buttons(contbuttons, 1)
                )
            bot.register_next_step_handler(message, report.period5)

    # Итоги дня
    def reportall1(message):
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        if message.text == '🌞 Сегодня':
            logging.info(f'Формирование отчета для {message.chat.id}')
            daterep = str(datetime.now().strftime("%d.%m.%Y"))
            report.rep(message, daterep, 1, 1, 1, 1, 1)
        elif message.text == '🗓️ Другой день':
            bot.send_message(
                message.chat.id,
                'Укажите дату в формате:\nПРИМЕР: 01.01.2023 или 01,01,2023',
                reply_markup = buttons.clearbuttons()
            )
            bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)
            bot.register_next_step_handler(message, report.reportallq)
        else:
            bot.send_message(
                message.chat.id,
                'Не верная команда',
                reply_markup = buttons.Buttons(['🌞 Сегодня', '🗓️ Другой день'])
            )
            bot.register_next_step_handler(message, report.reportall1)
    def reportallq(message):
        global ActiveUser
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        ActiveUser[message.chat.id]['repotherdate'] = message.text
        bot.send_message(
            message.chat.id,
            'Какие заявки показать?',
            reply_markup=buttons.Buttons(['Все', 'Только мои', 'У мастера'])
        )
        bot.register_next_step_handler(message, report.reportall2)
    def reportall2(message):
        global ActiveUser
        username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
        logging.info(f'{username} Отправил запрос - {message.text}')
        m1 = ActiveUser[message.chat.id]['repotherdate']
        m1 = m1.replace(' ', '.')
        m1 = m1.replace(',', '.')
        m = m1.split('.')
        if len(m[0]) == 2 and len(m[1]) == 2 and len(m[2]) == 4 and len(m) == 3:
            ActiveUser[message.chat.id]['daterep'] = m1
        daterep = str(ActiveUser[message.chat.id]['daterep'])
        print(daterep)
        if message.text.split()[0].isdigit():
            print('master')
            masterid = int(message.text.split()[0])
            mastername = str(db.get_record_by_id('Users', masterid)[2]) + ' ' + str(db.get_record_by_id('Users', masterid)[1])
            bot.send_message(
                message.chat.id,
                f'🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻\nОтчет за: {daterep}\nМастер: {mastername}',
                reply_markup=buttons.clearbuttons()
            )
            rept = db.select_table_with_filters('Tasks', {'master': masterid, 'status': 3}, ['done'], [daterep+' 00:00'], [daterep+' 23:59'])
            sendrep(message, rept)
            sendrepfile(message, rept)
            bot.send_message(
                message.chat.id,
                'Выберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.register_next_step_handler(message, MainMenu.Main2)
        elif message.text == 'Все':
            bot.send_message(
                message.chat.id,
                f'🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻\nОтчет за: {daterep}\nМастер: Все',
                reply_markup=buttons.clearbuttons()
            )
            rept = db.select_table_with_filters('Tasks', {'status': 3}, ['done'], [daterep+' 00:00'], [daterep+' 23:59'])
            sendrep(message, rept)
            sendrepfile(message, rept)
            bot.send_message(
                message.chat.id,
                'Выберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.register_next_step_handler(message, MainMenu.Main2)
        elif message.text == 'Только мои':
            bot.send_message(
                message.chat.id,
                f'🔻🔻🔻🔻🔻🔻🔻🔻🔻🔻\nОтчет за: {daterep}\nМастер: Я',
                reply_markup=buttons.clearbuttons()
            )
            rept = db.select_table_with_filters('Tasks', {'master': message.chat.id, 'status': 3}, ['done'], [daterep+' 00:00'], [daterep+' 23:59'])
            sendrep(message, rept)
            sendrepfile(message, rept)
            bot.send_message(
                message.chat.id,
                'Выберите операцию.',
                reply_markup=buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '🖨️ Обновить список техники', '📋 Мои заявки', '✏️ Редактировать контрагента', '📈 Отчеты', '🗺️ Карта', '📢 Написать всем'],3)
            )
            bot.register_next_step_handler(message, MainMenu.Main2)
        elif message.text == 'У мастера':
            users = db.select_table('Users')
            masters = []
            masters.append('↩️ Назад')
            for user in users:
                line = str(user[0]) + ' ' + str(user[2]) + ' ' + str(user[1])
                masters.append(line)
            bot.send_message(
                message.chat.id,
                'Выберите мастера.',
                reply_markup=buttons.Buttons(masters, 1)
            )
            bot.register_next_step_handler(message, report.reportall2)
        else:
            bot.send_message(
                message.chat.id,
                'Не верная команда!\nКакие заявки показать?',
                reply_markup=buttons.Buttons(['Все', 'Только мои', 'У мастера'])
            )
            bot.register_next_step_handler(message, report.reportall2)

@bot.message_handler(content_types=['text', 'location'])
# формирование гугл ссылки на карты по локации для адреса компании
def CADR1(message):
    username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
    logging.info(f'{username} Отправил запрос - {message.text}')
    global ActiveUser
    if message.content_type == 'location':
        lon, lat = message.location.longitude, message.location.latitude
        url = f'GOOGLE: https://www.google.com/maps/search/?api=1&query={lat},{lon}'
        ActiveUser[message.chat.id]['contnew'][2] = url
    else:
        ActiveUser[message.chat.id]['contnew'][2] = message.text
    editcontragent(message)
    bot.register_next_step_handler(message, editcont.ec2)
# Формирование гугл ссылки на карты по локации для адреса при добавлении нового контрагента
def NeContr4(message):
    username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
    logging.info(f'{username} Отправил запрос - {message.text}')
    global ActiveUser
    if message.content_type == 'location':
        lon, lat = message.location.longitude, message.location.latitude
        url = f'GOOGLE: https://www.google.com/maps/search/?api=1&query={lat},{lon}\nAPPLE: http://maps.apple.com/maps?ll={lat},{lon}'
        ActiveUser[message.chat.id]['cadr'] = url
    else:
        ActiveUser[message.chat.id]['cadr'] = message.text
    bot.send_message(
        message.chat.id,
        'Кто подал заявку? Укажите имя контактного лица.',
        reply_markup=buttons.clearbuttons()
    )
    bot.register_next_step_handler(message, NeContr5)
def NeContr5(message):
    username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
    logging.info(f'{username} Отправил запрос - {message.text}')
    global ActiveUser
    if ActiveUser[message.chat.id]['ds'] == 3:
        if message.content_type == 'location':
            lon, lat = message.location.longitude, message.location.latitude
            url = f'GOOGLE: https://www.google.com/maps/search/?api=1&query={lat},{lon}\nAPPLE: http://maps.apple.com/maps?ll={lat},{lon}'
            ActiveUser[message.chat.id]['cadr'] = url
        else:
            ActiveUser[message.chat.id]['cadr'] = message.text
    else:
        ActiveUser[message.chat.id]['cperson'] = message.text
    bot.send_message(
        message.chat.id,
        'Укажите контактный номер телефона для связи с клиентом.',
        reply_markup=buttons.clearbuttons()
    )
    bot.register_next_step_handler(message, NewTask.NeContr6)
# Добавление новой локации в список локаций в редактировании контрагента
def newlocation(message):
    username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
    logging.info(f'{username} Отправил запрос - {message.text}')
    global ActiveUser
    if message.content_type == 'location':
        ActiveUser[message.chat.id]['lon'], ActiveUser[message.chat.id]['lat'] = message.location.longitude, message.location.latitude
        bot.send_message(
            message.chat.id,
            'Укажите название локации\nНАПРИМЕР:\nФилиал чиланзар или головной офис',
            reply_markup=buttons.clearbuttons()
        )
        bot.register_next_step_handler(message, newlocation1)
    else:
        bot.send_message(
            message.chat.id,
            'Вы должны были отправить локацию.\nОтправьте локацию.',
            reply_markup=buttons.clearbuttons
        )
        bot.register_next_step_handler(message, newlocation)
def newlocation1(message):
    username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
    logging.info(f'{username} Отправил запрос - {message.text}')
    global ActiveUser
    ActiveUser[message.chat.id]['locationname'] = message.text
    db.insert_record(
        'Locations',
        [
            None,
            ActiveUser[message.chat.id]['inn'],
            ActiveUser[message.chat.id]['locationname'],
            ActiveUser[message.chat.id]['lat'],
            ActiveUser[message.chat.id]['lon']
        ]
    )
    bot.send_message(
        message.chat.id,
        'Локация сохранена.',
        reply_markup=buttons.clearbuttons()
    )
    locations = db.select_table_with_filters('Locations', {'inn': ActiveUser[message.chat.id]['inn']})
    buttonsloc = []
    buttonsloc.append('🆕 Добавить')
    try:
        for location in locations:
            buttonsloc.append(str(location[0]) + ' ' + str(location[2]))
    except Exception as e:
        logging.error(e)
        time.sleep(5)
    buttonsloc.append('↩️ Назад')
    if len(locations) > 2:
        bot.send_message(
            message.chat.id,
            'Это все локации выберите ту которую хотите изменить.',
            reply_markup=buttons.Buttons(buttonsloc, 2)
        )
    else:
        bot.send_message(
            message.chat.id,
            'Для выбранного контрагента не указаны локации.',
            reply_markup=buttons.Buttons(buttonsloc, 2)
        )
    bot.register_next_step_handler(message, editcont.locations1)
# Изменение локаии контрагента
def editcontlocation1(message):
    username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
    logging.info(f'{username} Отправил запрос - {message.text}')
    global ActiveUser
    if message.content_type == 'location':
        lon, lat = message.location.longitude, message.location.latitude
        db.update_records(
            'Locations',
            [
                'lon',
                'lat'
            ],
            [
                lon,
                lat
            ],
            'id',
            ActiveUser[message.chat.id]['curlocation']
        )
        bot.send_message(
            message.chat.id,
            'Локация изменена.\nЧто вы хотите изменить?',
            reply_markup=buttons.Buttons(['Изменить локацию', 'Изменить название', '🗑️ Удалить', '🚫 Отмена'], 3)
        )
        bot.register_next_step_handler(message, editcont.locations2)
    else:
        bot.send_message(
            message.chat.id,
            'Вы должны были отправить локацию.\nОтправьте локацию.',
            reply_markup=buttons.clearbuttons
        )
        bot.register_next_step_handler(message, newlocation)
# Добавление локации филиала в новой заявке
def newlocationintask1(message):
    username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
    logging.info(f'{username} Отправил запрос - {message.text}')
    global ActiveUser
    if message.content_type == 'location':
        ActiveUser[message.chat.id]['lon'], ActiveUser[message.chat.id]['lat'] = message.location.longitude, message.location.latitude
        bot.send_message(
            message.chat.id,
            'Укажите название локации\nНАПРИМЕР:\nФилиал чиланзар или головной офис',
            reply_markup=buttons.clearbuttons()
        )
        bot.register_next_step_handler(message, newlocationintask2)
    else:
        bot.send_message(
            message.chat.id,
            'Вы должны были отправить локацию.\nОтправьте локацию.',
            reply_markup=buttons.clearbuttons
        )
        bot.register_next_step_handler(message, newlocationintask1)
def newlocationintask2(message):
    username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
    logging.info(f'{username} Отправил запрос - {message.text}')
    global ActiveUser
    ActiveUser[message.chat.id]['locationname'] = message.text
    db.insert_record(
        'Locations',
        [
            None,
            ActiveUser[message.chat.id]['inn'],
            ActiveUser[message.chat.id]['locationname'],
            ActiveUser[message.chat.id]['lat'],
            ActiveUser[message.chat.id]['lon']
        ]
    )
    locations = db.select_table_with_filters('Locations', {'inn': ActiveUser[message.chat.id]['inn']})
    clocations = ['⏩ Пропустить']
    if len(locations) > 0:
        for i in locations:
            line = str(i[0]) + ' ' + str(i[2])
            clocations.append(line)
        clocations.append('🆕 Добавить филиал')
        bot.send_message(
            message.chat.id,
            'Выберите филиал',
            reply_markup=buttons.Buttons(clocations)
        )
    bot.register_next_step_handler(message, NewTask.ntlocation2)
# Добавление локации филиала в акнивной заявке
def tnl1(message):
    username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
    logging.info(f'{username} Отправил запрос - {message.text}')
    global ActiveUser
    if message.content_type == 'location':
        lon, lat = message.location.longitude, message.location.latitude
        ActiveUser[message.chat.id]['lon'] = lon
        ActiveUser[message.chat.id]['lat'] = lat
        bot.send_message(
            message.chat.id,
            'Укажите название локации\nНАПРИМЕР:\nФилиал чиланзар или головной офис',
        )
        bot.register_next_step_handler(message, tnl2)
    else:
        bot.send_message(
            message.chat.id,
            'Вы должны были отправить локацию.\nОтправьте локацию.',
            reply_markup=buttons.clearbuttons
        )
        bot.register_next_step_handler(message, tnl1)
def tnl2(message):
    username = db.get_record_by_id('Users', message.chat.id)[2] + ' ' + db.get_record_by_id('Users', message.chat.id)[1]
    logging.info(f'{username} Отправил запрос - {message.text}')
    global ActiveUser
    ActiveUser[message.chat.id]['locationname'] = message.text
    ActiveUser[message.chat.id]['inn'] = db.get_record_by_id('Tasks', ActiveUser[message.chat.id]['task'])[3]
    db.insert_record(
        'Locations',
        [
            None,
            ActiveUser[message.chat.id]['inn'],
            ActiveUser[message.chat.id]['locationname'],
            ActiveUser[message.chat.id]['lat'],
            ActiveUser[message.chat.id]['lon']
        ]
    )
    inn = db.get_record_by_id('Tasks', ActiveUser[message.chat.id]['task'])[3]
    locations = db.select_table_with_filters('Locations', {'inn': inn})
    buttonsloc = []
    buttonsloc.append('🆕 Добавить филиал')
    if len(locations) > 0:
        for location in locations:
            buttonsloc.append(str(location[0]) + ' ' + str(location[2]))
    buttonsloc.append('🚫 Отмена')
    if len(locations) > 0:
        bot.send_message(
            message.chat.id,
            'Выберите филиал...',
            reply_markup=buttons.Buttons(buttonsloc, 2)
        )
    bot.register_next_step_handler(message, Task.locations2)


@bot.callback_query_handler(func=lambda call: True)
# реакция на инлайновые кнопки
def callback_handler(call):
    global ActiveUser, sendedmessages, continue_polling
    continue_polling = False
    if call.data.split()[0] == 'tasklist':# Подробности заявки
        status = db.get_record_by_id('Tasks', int(call.data.split()[1]))
        if status[11] == 1 or status[11] == 5:
            markdownt = buttons.Buttons(['👍 Принять', '🖊️ Дополнить', '📎 Назначить', '🤵 Изменить контрагента', '✏️ Изменить текст заявки', '📍 Локация', '🚫 Отменить заявку', '↩️ Назад'])
        elif status[11] == 2 or status[11] == 6:
            markdownt = buttons.Buttons(['✅ Выполнено', '🖊️ Дополнить', '🙅‍♂️ Отказаться от заявки', '📎 Переназначить', '🤵 Изменить контрагента', '✏️ Изменить текст заявки', '📍 Локация', '🚫 Отменить заявку', '↩️ Назад'], 3)
        else:
            markdownt = buttons.Buttons(['📝 Новая заявка', '🔃 Обновить список заявок', '📋 Мои заявки', '✏️ Редактировать контрагента', '✏️ Изменить текст заявки', '📍 Локация', '📈 Отчеты', '📢 Написать всем'],3)
        ActiveUser[call.from_user.id]['sentmes'] = bot.send_message(
            call.from_user.id,
            functions.curtask(call.data.split()[1]),
            reply_markup=markdownt
        )
        ActiveUser[call.from_user.id]['task'] = call.data.split()[1]
        bot.register_next_step_handler(call.message, Task.task1)
    elif call.data.split()[0] == 'confirm':# Принятие заявки
        if db.get_record_by_id('Tasks', call.data.split()[1])[11] == 5:
            stat = 6
        else:
            stat = 2
        if db.get_record_by_id('Tasks', call.data.split()[1])[11] != 1:
            bot.send_message(
                call.from_user.id,
                "Вы не можете принять эту заявку! ее уже принял " + db.get_record_by_id('Users', db.get_record_by_id('Tasks', ActiveUser[call.from_user.id]['task'])[6])[2] + ' ' + db.get_record_by_id('Users', db.get_record_by_id('Tasks', ActiveUser[call.from_user.id]['task'])[6])[1]
            )
        else:
            db.update_records(
                'Tasks',
                [
                    'confirmed',
                    'master',
                    'status'
                ], [
                    datetime.now().strftime("%d.%m.%Y %H:%M"),
                    call.from_user.id,
                    stat
                ],
                'id',
                call.data.split()[1]
            )
            bot.send_message(
                call.from_user.id,
                "Вы приняли заявку..."
            )
            sendtoall(str(db.get_record_by_id('Users', call.from_user.id)[2]) + ' ' + str(db.get_record_by_id('Users', call.from_user.id)[1]) + '\nПринял заявку:\n\n' + functions.curtask(call.data.split()[1]), '', call.from_user.id)
            bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
            continue_polling = True
            # bot.register_next_step_handler(call.message, MainMenu.Main2)
            for line in sendedmessages:
                bot.delete_message(line[0], line[1])
    elif call.data.split()[0] == 'set':# Назначение мастера
        users = db.select_table('Users')
        bot.send_message(
            call.from_user.id,
            'Выберите мастера...',
            reply_markup=buttons.Buttons(functions.listgen(users, [0, 1, 2], 3), 1)
        )
        bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
        ActiveUser[call.from_user.id]['task'] = call.data.split()[1]
        bot.register_next_step_handler(call.message, Task.task4)
        
# Запуск бота
if __name__ == '__main__':
    sendtoall('‼️‼️‼️Сервер бота был перезагружен...‼️‼️‼️\nНажмите кнопку "/start"', buttons.Buttons(['/start']), 0, 0, True)
    thread = threading.Thread(target=asyncio.run, args=(main(),))
    thread.start()
    if continue_polling is True:
        # bot.polling(none_stop=True, interval=0)
        while True:
            try:
                bot.polling(none_stop=True, interval=0)
                logging.info('запуск пула')
                logging.info()
            except Exception as e:
                logging.error(e)
                time.sleep(5)